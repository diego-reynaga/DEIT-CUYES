#!/usr/bin/env python3
"""
Servidor completo del Sistema de Gestión de Cuyes - INIA Andahuaylas
"""
import os
from datetime import datetime
from dotenv import load_dotenv
load_dotenv()

# ========== TEMPLATES HTML ==========
HOME_TEMPLATE = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Sistema de Gestión de Cuyes - INIA</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
</head>
<body class="bg-light">
    <div class="container mt-5">
        <div class="row justify-content-center">
            <div class="col-md-8 text-center">
                <h1 class="display-4 text-primary mb-4">
                    <i class="fas fa-seedling"></i> Sistema de Gestión de Cuyes
                </h1>
                <h2 class="text-secondary mb-4">INIA Andahuaylas</h2>
                <p class="lead mb-4">Sistema integral para la gestión, control y venta de cuyes</p>
                
                <div class="row mt-5">
                    <div class="col-md-6 mb-3">
                        <div class="card h-100">
                            <div class="card-body">
                                <i class="fas fa-users fa-3x text-primary mb-3"></i>
                                <h5>Gestión de Cuyes</h5>
                                <p>Control completo del inventario de animales</p>
                            </div>
                        </div>
                    </div>
                    <div class="col-md-6 mb-3">
                        <div class="card h-100">
                            <div class="card-body">
                                <i class="fas fa-chart-line fa-3x text-success mb-3"></i>
                                <h5>Sistema de Ventas</h5>
                                <p>Gestión de ventas y clientes</p>
                            </div>
                        </div>
                    </div>
                </div>
                
                <div class="mt-4">
                    <a href="/login" class="btn btn-primary btn-lg me-2">
                        <i class="fas fa-sign-in-alt"></i> Iniciar Sesión
                    </a>
                    <a href="/registro" class="btn btn-success btn-lg me-2">
                        <i class="fas fa-user-plus"></i> Crear Cuenta
                    </a>
                    <a href="/catalogo" class="btn btn-outline-info btn-lg">
                        <i class="fas fa-store"></i> Ver Catálogo
                    </a>
                </div>
            </div>
        </div>
    </div>
</body>
</html>
"""

LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Iniciar Sesión - Sistema de Cuyes</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
</head>
<body class="bg-light">
    <div class="container">
        <div class="row justify-content-center mt-5">
            <div class="col-md-6">
                <div class="card shadow">
                    <div class="card-header bg-primary text-white text-center">
                        <h4><i class="fas fa-user-lock"></i> Iniciar Sesión</h4>
                    </div>
                    <div class="card-body">
                        {% with messages = get_flashed_messages(with_categories=true) %}
                        {% if messages %}
                        {% for category, message in messages %}
                        <div class="alert alert-{{ 'danger' if category == 'danger' else 'info' }}">{{ message }}</div>
                        {% endfor %}
                        {% endif %}
                        {% endwith %}
                        
                        <form method="POST">
                            {{ form.hidden_tag() }}
                            <div class="mb-3">
                                {{ form.email.label(class="form-label") }}
                                {{ form.email(class="form-control") }}
                            </div>
                            <div class="mb-3">
                                {{ form.password.label(class="form-label") }}
                                {{ form.password(class="form-control") }}
                            </div>
                            <button type="submit" class="btn btn-primary w-100">
                                <i class="fas fa-sign-in-alt"></i> Entrar
                            </button>
                        </form>
                        
                    </div>
                    <div class="card-footer text-center">
                        <a href="/" class="btn btn-outline-secondary me-2">
                            <i class="fas fa-arrow-left"></i> Volver al Inicio
                        </a>
                        <a href="/registro" class="btn btn-outline-success">
                            <i class="fas fa-user-plus"></i> Crear Cuenta
                        </a>
                    </div>
                </div>
            </div>
        </div>
    </div>
</body>
</html>
"""

DASHBOARD_TEMPLATE = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Dashboard - Sistema de Cuyes</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
</head>
<body class="bg-light">
    <nav class="navbar navbar-expand-lg navbar-dark bg-primary">
        <div class="container-fluid">
            <a class="navbar-brand" href="/">
                <i class="fas fa-seedling"></i> Sistema de Cuyes
            </a>
            <div class="navbar-nav ms-auto">
                <span class="navbar-text me-3">Bienvenido, {{ user.nombre }}!</span>
                <a class="nav-link" href="/logout">
                    <i class="fas fa-sign-out-alt"></i> Salir
                </a>
            </div>
        </div>
    </nav>
    
    <div class="container mt-4">
        <h2>Dashboard Principal</h2>
        
        <!-- Estadísticas -->
        <div class="row mb-4">
            <div class="col-md-3">
                <div class="card bg-primary text-white">
                    <div class="card-body">
                        <div class="d-flex justify-content-between">
                            <div>
                                <h5>Cuyes</h5>
                                <h3>{{ stats.total_cuyes }}</h3>
                            </div>
                            <i class="fas fa-users fa-2x"></i>
                        </div>
                    </div>
                </div>
            </div>
            <div class="col-md-3">
                <div class="card bg-success text-white">
                    <div class="card-body">
                        <div class="d-flex justify-content-between">
                            <div>
                                <h5>Razas</h5>
                                <h3>{{ stats.total_razas }}</h3>
                            </div>
                            <i class="fas fa-dna fa-2x"></i>
                        </div>
                    </div>
                </div>
            </div>
            <div class="col-md-3">
                <div class="card bg-warning text-white">
                    <div class="card-body">
                        <div class="d-flex justify-content-between">
                            <div>
                                <h5>Pozas</h5>
                                <h3>{{ stats.total_pozas }}</h3>
                            </div>
                            <i class="fas fa-home fa-2x"></i>
                        </div>
                    </div>
                </div>
            </div>
            <div class="col-md-3">
                <div class="card bg-info text-white">
                    <div class="card-body">
                        <div class="d-flex justify-content-between">
                            <div>
                                <h5>Ventas</h5>
                                <h3>{{ stats.total_ventas }}</h3>
                            </div>
                            <i class="fas fa-shopping-cart fa-2x"></i>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        
        <!-- Menú Principal -->
        <div class="row">
            <div class="col-md-4 mb-3">
                <div class="card h-100">
                    <div class="card-body text-center">
                        <i class="fas fa-users fa-3x text-primary mb-3"></i>
                        <h5>Gestión de Cuyes</h5>
                        <p>Administrar inventario de cuyes</p>
                        <a href="/cuyes" class="btn btn-primary">Acceder</a>
                    </div>
                </div>
            </div>
            <div class="col-md-4 mb-3">
                <div class="card h-100">
                    <div class="card-body text-center">
                        <i class="fas fa-dna fa-3x text-success mb-3"></i>
                        <h5>Razas</h5>
                        <p>Gestionar razas de cuyes</p>
                        <a href="/razas" class="btn btn-success">Acceder</a>
                    </div>
                </div>
            </div>
            <div class="col-md-4 mb-3">
                <div class="card h-100">
                    <div class="card-body text-center">
                        <i class="fas fa-home fa-3x text-warning mb-3"></i>
                        <h5>Pozas</h5>
                        <p>Administrar pozas y espacios</p>
                        <a href="/pozas" class="btn btn-warning">Acceder</a>
                    </div>
                </div>
            </div>
            <div class="col-md-4 mb-3">
                <div class="card h-100">
                    <div class="card-body text-center">
                        <i class="fas fa-shopping-cart fa-3x text-info mb-3"></i>
                        <h5>Ventas</h5>
                        <p>Sistema de ventas</p>
                        <a href="/ventas" class="btn btn-info">Acceder</a>
                    </div>
                </div>
            </div>
            <div class="col-md-4 mb-3">
                <div class="card h-100">
                    <div class="card-body text-center">
                        <i class="fas fa-chart-bar fa-3x text-secondary mb-3"></i>
                        <h5>Reportes</h5>
                        <p>Generar reportes y estadísticas</p>
                        <a href="/reportes" class="btn btn-secondary">Acceder</a>
                    </div>
                </div>
            </div>
        </div>
    </div>
</body>
</html>
"""

try:
    print("=== SISTEMA DE GESTIÓN DE CUYES - INIA ANDAHUAYLAS ===")
    print("Iniciando servidor completo...")
    
    import os
    from datetime import datetime, date
    from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, render_template_string, make_response, send_file
    from flask_sqlalchemy import SQLAlchemy
    from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
    from werkzeug.security import generate_password_hash, check_password_hash
    from flask_wtf import FlaskForm
    from wtforms import StringField, PasswordField, SelectField, FloatField, TextAreaField, DateField, BooleanField, IntegerField
    from wtforms.validators import DataRequired, Email, Length
    from templates_extra import CUYES_TEMPLATE, FORM_TEMPLATE, BASE_TEMPLATE
    
    # Importaciones para generar PDF y Excel
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    import pandas as pd
    import tempfile
    import io
    
    # Crear aplicación Flask
    app = Flask(__name__)
    app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-key-cuyes-2024')
    app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'mysql+pymysql://root:12345678@localhost/cuyes_db')
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    
    # Inicializar extensiones
    db = SQLAlchemy(app)
    login_manager = LoginManager(app)
    login_manager.login_view = 'login'
    login_manager.login_message = 'Inicia sesión para acceder'
    
    # ========== MODELOS ==========
    class Usuario(UserMixin, db.Model):
        __tablename__ = 'usuarios'
        id = db.Column(db.Integer, primary_key=True)
        nombre = db.Column(db.String(100), nullable=False)
        apellido = db.Column(db.String(100), nullable=False)
        email = db.Column(db.String(120), unique=True, nullable=False)
        password_hash = db.Column(db.String(255), nullable=False)
        rol = db.Column(db.String(20), default='cliente')
        activo = db.Column(db.Boolean, default=True)
        telefono = db.Column(db.String(20))
        direccion = db.Column(db.Text)
        fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)
        
        def check_password(self, password):
            return check_password_hash(self.password_hash, password)
        
        def is_admin(self):
            return self.rol == 'admin'
        
        def is_empleado(self):
            return self.rol == 'empleado'
        
        def is_cliente(self):
            return self.rol == 'cliente'
        
        @property
        def nombre_completo(self):
            return f"{self.nombre} {self.apellido}"
    
    class Raza(db.Model):
        __tablename__ = 'razas'
        id = db.Column(db.Integer, primary_key=True)
        nombre = db.Column(db.String(100), unique=True, nullable=False)
        descripcion = db.Column(db.Text)
    
    class Poza(db.Model):
        __tablename__ = 'pozas'
        id = db.Column(db.Integer, primary_key=True)
        codigo = db.Column(db.String(20), unique=True, nullable=False)
        nombre = db.Column(db.String(100))  # Nombre descriptivo
        tipo = db.Column(db.String(50), nullable=False)
        capacidad = db.Column(db.Integer, nullable=False)
        ocupacion_actual = db.Column(db.Integer, default=0)
        activa = db.Column(db.Boolean, default=True)
        descripcion = db.Column(db.Text)
        
        def __init__(self, **kwargs):
            super(Poza, self).__init__(**kwargs)
            if not self.nombre:
                self.nombre = self.codigo
    
    class Cuy(db.Model):
        __tablename__ = 'cuyes'
        id = db.Column(db.Integer, primary_key=True)
        codigo = db.Column(db.String(20), unique=True, nullable=False)
        numero = db.Column(db.String(20), unique=True)  # Número identificador
        sexo = db.Column(db.String(10), nullable=False)
        genero = db.Column(db.String(10))  # Alias para sexo
        raza = db.Column(db.String(50))  # Nombre de la raza directamente
        fecha_nacimiento = db.Column(db.Date)
        peso_actual = db.Column(db.Float(precision=2))
        peso = db.Column(db.Float(precision=2))  # Alias para peso_actual
        estado = db.Column(db.String(20), default='sano')
        precio_venta = db.Column(db.Float(precision=2))
        raza_id = db.Column(db.Integer, db.ForeignKey('razas.id'))
        poza_id = db.Column(db.Integer, db.ForeignKey('pozas.id'))
        observaciones = db.Column(db.Text)
        fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)
        
        # Relaciones
        raza_obj = db.relationship('Raza', backref='cuyes_rel')
        poza = db.relationship('Poza', backref='cuyes')
        
        def __init__(self, **kwargs):
            super(Cuy, self).__init__(**kwargs)
            if not self.numero:
                self.numero = self.codigo
            if not self.genero:
                self.genero = self.sexo
            if not self.peso:
                self.peso = self.peso_actual
            if self.raza_obj and not self.raza:
                self.raza = self.raza_obj.nombre
    
    class Venta(db.Model):
        __tablename__ = 'ventas'
        id = db.Column(db.Integer, primary_key=True)
        cliente_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'))
        cuy_id = db.Column(db.Integer, db.ForeignKey('cuyes.id'))
        fecha_venta = db.Column(db.Date, nullable=False)
        precio = db.Column(db.Float(precision=2), nullable=False)
        total = db.Column(db.Float(precision=2))  # Alias para precio
        cantidad = db.Column(db.Integer, default=1)
        peso_venta = db.Column(db.Float(precision=2))
        metodo_pago = db.Column(db.String(20), default='transferencia')
        estado = db.Column(db.String(20), default='pendiente')
        nombre_cliente = db.Column(db.String(200))
        email_cliente = db.Column(db.String(120))
        telefono_cliente = db.Column(db.String(20))
        observaciones = db.Column(db.Text)
        vendido_por = db.Column(db.Integer, db.ForeignKey('usuarios.id'))
        fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)
        
        # Relaciones
        cliente = db.relationship('Usuario', foreign_keys=[cliente_id], backref='compras')
        vendedor = db.relationship('Usuario', foreign_keys=[vendido_por])
        cuy = db.relationship('Cuy', backref='venta')
        
        def __init__(self, **kwargs):
            super(Venta, self).__init__(**kwargs)
            if not self.total:
                self.total = self.precio
            if self.cliente and not self.nombre_cliente:
                self.nombre_cliente = self.cliente.nombre_completo
            if self.cliente and not self.email_cliente:
                self.email_cliente = self.cliente.email
    
    class Control(db.Model):
        __tablename__ = 'controles'
        id = db.Column(db.Integer, primary_key=True)
        cuy_id = db.Column(db.Integer, db.ForeignKey('cuyes.id'), nullable=False)
        fecha_control = db.Column(db.Date, nullable=False)
        peso = db.Column(db.Float(precision=2))
        temperatura = db.Column(db.Float(precision=2))
        estado_salud = db.Column(db.String(50))
        observaciones = db.Column(db.Text)
        usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'))
        fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)
        
        # Relaciones
        cuy = db.relationship('Cuy', backref='controles')
        usuario = db.relationship('Usuario', foreign_keys=[usuario_id])
    
    class Tratamiento(db.Model):
        __tablename__ = 'tratamientos'
        id = db.Column(db.Integer, primary_key=True)
        cuy_id = db.Column(db.Integer, db.ForeignKey('cuyes.id'), nullable=False)
        fecha_inicio = db.Column(db.Date, nullable=False)
        fecha_fin = db.Column(db.Date)
        diagnostico = db.Column(db.String(200), nullable=False)
        medicamento = db.Column(db.String(100))
        dosis = db.Column(db.String(50))
        frecuencia = db.Column(db.String(50))
        duracion_dias = db.Column(db.Integer)
        estado = db.Column(db.String(20), default='en_curso')
        observaciones = db.Column(db.Text)
        usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'))
        fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)
        
        # Relaciones
        cuy = db.relationship('Cuy', backref='tratamientos')
        usuario = db.relationship('Usuario', foreign_keys=[usuario_id])
    
    class Comentario(db.Model):
        __tablename__ = 'comentarios'
        id = db.Column(db.Integer, primary_key=True)
        cuy_id = db.Column(db.Integer, db.ForeignKey('cuyes.id'), nullable=False)
        usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
        venta_id = db.Column(db.Integer, db.ForeignKey('ventas.id'))
        calificacion = db.Column(db.Integer, nullable=False)  # 1-5 estrellas
        comentario = db.Column(db.Text)
        aprobado = db.Column(db.Boolean, default=False)
        fecha_comentario = db.Column(db.DateTime, default=datetime.utcnow)
        
        # Relaciones
        cuy = db.relationship('Cuy', backref='comentarios')
        usuario = db.relationship('Usuario')
        venta = db.relationship('Venta')
    
    @login_manager.user_loader
    def load_user(user_id):
        return Usuario.query.get(int(user_id))
    
    # ========== FORMULARIOS ==========
    class LoginForm(FlaskForm):
        email = StringField('Email', validators=[DataRequired(), Email()])
        password = PasswordField('Contraseña', validators=[DataRequired()])
    
    class CuyForm(FlaskForm):
        codigo = StringField('Código', validators=[DataRequired()])
        sexo = SelectField('Sexo', choices=[('macho', 'Macho'), ('hembra', 'Hembra')])
        peso_actual = FloatField('Peso (kg)', validators=[DataRequired()])
        precio_venta = FloatField('Precio de Venta (S/.)')
        fecha_nacimiento = DateField('Fecha de Nacimiento', default=datetime.today)
        raza_id = SelectField('Raza', coerce=int)
        poza_id = SelectField('Poza', coerce=int)
        observaciones = TextAreaField('Observaciones')
    
    class RazaForm(FlaskForm):
        nombre = StringField('Nombre', validators=[DataRequired()])
        descripcion = TextAreaField('Descripción')
    
    class PozaForm(FlaskForm):
        codigo = StringField('Código', validators=[DataRequired()])
        tipo = SelectField('Tipo', choices=[('reproductores', 'Reproductores'), ('lactantes', 'Lactantes'), ('recria', 'Recría'), ('engorde', 'Engorde')])
        capacidad = StringField('Capacidad', validators=[DataRequired()])
    
    class VentaForm(FlaskForm):
        cuy_id = SelectField('Cuy', coerce=int)
        precio = FloatField('Precio (S/.)', validators=[DataRequired()])
        peso_venta = FloatField('Peso en Venta (kg)')
        metodo_pago = SelectField('Método de Pago', choices=[
            ('efectivo', 'Efectivo'),
            ('transferencia', 'Transferencia'),
            ('tarjeta', 'Tarjeta')
        ])
        cliente_nombre = StringField('Nombre del Cliente')
        observaciones = TextAreaField('Observaciones')
    
    class ControlForm(FlaskForm):
        cuy_id = SelectField('Cuy', coerce=int)
        fecha_control = DateField('Fecha de Control', validators=[DataRequired()])
        peso = FloatField('Peso (kg)', validators=[DataRequired()])
        temperatura = FloatField('Temperatura (°C)')
        estado_salud = SelectField('Estado de Salud', choices=[
            ('excelente', 'Excelente'),
            ('buena', 'Buena'),
            ('regular', 'Regular'),
            ('mala', 'Mala')
        ])
        observaciones = TextAreaField('Observaciones')
    
    class TratamientoForm(FlaskForm):
        cuy_id = SelectField('Cuy', coerce=int)
        fecha_inicio = DateField('Fecha de Inicio', validators=[DataRequired()])
        fecha_fin = DateField('Fecha de Fin')
        diagnostico = StringField('Diagnóstico', validators=[DataRequired()])
        medicamento = StringField('Medicamento')
        dosis = StringField('Dosis')
        frecuencia = SelectField('Frecuencia', choices=[
            ('diario', 'Diario'),
            ('cada_12h', 'Cada 12 horas'),
            ('cada_8h', 'Cada 8 horas'),
            ('semanal', 'Semanal')
        ])
        duracion_dias = IntegerField('Duración (días)')
        observaciones = TextAreaField('Observaciones')
        completado = BooleanField('Tratamiento Completado')
    
    class ComentarioForm(FlaskForm):
        calificacion = SelectField('Calificación', choices=[
            ('5', '⭐⭐⭐⭐⭐ Excelente'),
            ('4', '⭐⭐⭐⭐ Muy Bueno'),
            ('3', '⭐⭐⭐ Bueno'),
            ('2', '⭐⭐ Regular'),
            ('1', '⭐ Malo')
        ], coerce=int, validators=[DataRequired()])
        comentario = TextAreaField('Comentario', validators=[DataRequired()])
    
    # ========== FUNCIONES AUXILIARES ==========
    def actualizar_ocupacion_pozas():
        """Actualiza la ocupación actual de todas las pozas basándose en el conteo real de cuyes"""
        pozas = Poza.query.all()
        for poza in pozas:
            cuyes_actuales = Cuy.query.filter_by(poza_id=poza.id, estado='sano').count()
            poza.ocupacion_actual = cuyes_actuales
        db.session.commit()

    # ========== RUTAS ==========
    @app.route('/')
    def index():
        return render_template_string(HOME_TEMPLATE)
    
    @app.route('/login', methods=['GET', 'POST'])
    def login():
        form = LoginForm()
        if form.validate_on_submit():
            user = Usuario.query.filter_by(email=form.email.data).first()
            if user and user.check_password(form.password.data) and user.activo:
                login_user(user)
                return redirect(url_for('dashboard'))
            else:
                flash('Credenciales incorrectas', 'danger')
        
        return render_template_string(LOGIN_TEMPLATE, form=form)
    
    # ========== REGISTRO PÚBLICO PARA CLIENTES ==========
    @app.route('/registro', methods=['GET', 'POST'])
    def registro():
        if request.method == 'POST':
            # Validar que el email no exista
            email = request.form.get('email')
            if Usuario.query.filter_by(email=email).first():
                flash('El email ya está registrado. Intenta con otro email.', 'danger')
                return redirect(url_for('registro'))
            
            # Crear nuevo usuario cliente
            usuario = Usuario(
                nombre=request.form.get('nombre'),
                apellido=request.form.get('apellido'),
                email=email,
                password_hash=generate_password_hash(request.form.get('password')),
                telefono=request.form.get('telefono', ''),
                direccion=request.form.get('direccion', ''),
                rol='cliente',  # Automáticamente asignado como cliente
                activo=True
            )
            
            db.session.add(usuario)
            db.session.commit()
            
            flash('¡Registro exitoso! Ya puedes iniciar sesión.', 'success')
            return redirect(url_for('login'))
        
        registro_content = """
        <div class="row justify-content-center">
            <div class="col-md-6">
                <div class="card">
                    <div class="card-header bg-success text-white text-center">
                        <h4><i class="fas fa-user-plus"></i> Registro de Cliente</h4>
                        <p class="mb-0">Únete a la comunidad INIA</p>
                    </div>
                    <div class="card-body">
                        <form method="POST">
                            <div class="row">
                                <div class="col-md-6 mb-3">
                                    <label for="nombre" class="form-label">Nombre *</label>
                                    <input type="text" class="form-control" id="nombre" name="nombre" required>
                                </div>
                                <div class="col-md-6 mb-3">
                                    <label for="apellido" class="form-label">Apellido *</label>
                                    <input type="text" class="form-control" id="apellido" name="apellido" required>
                                </div>
                            </div>
                            
                            <div class="mb-3">
                                <label for="email" class="form-label">Email *</label>
                                <input type="email" class="form-control" id="email" name="email" required>
                                <div class="form-text">Usa este email para iniciar sesión</div>
                            </div>
                            
                            <div class="mb-3">
                                <label for="password" class="form-label">Contraseña *</label>
                                <input type="password" class="form-control" id="password" name="password" required minlength="6">
                                <div class="form-text">Mínimo 6 caracteres</div>
                            </div>
                            
                            <div class="mb-3">
                                <label for="telefono" class="form-label">Teléfono</label>
                                <input type="tel" class="form-control" id="telefono" name="telefono" placeholder="Opcional">
                            </div>
                            
                            <div class="mb-3">
                                <label for="direccion" class="form-label">Dirección</label>
                                <textarea class="form-control" id="direccion" name="direccion" rows="2" placeholder="Opcional"></textarea>
                            </div>
                            
                            <div class="alert alert-info">
                                <small><i class="fas fa-info-circle"></i> 
                                Al registrarte, obtienes acceso al catálogo de cuyes INIA y podrás realizar compras.</small>
                            </div>
                            
                            <div class="d-grid gap-2">
                                <button type="submit" class="btn btn-success btn-lg">
                                    <i class="fas fa-user-plus"></i> Crear Cuenta
                                </button>
                                <a href="/login" class="btn btn-outline-secondary">
                                    <i class="fas fa-sign-in-alt"></i> Ya tengo cuenta
                                </a>
                                <a href="/" class="btn btn-link">
                                    <i class="fas fa-home"></i> Volver al inicio
                                </a>
                            </div>
                        </form>
                    </div>
                </div>
            </div>
        </div>
        """
        
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', registro_content))
    
    @app.route('/dashboard')
    @login_required
    def dashboard():
        # Obtener estadísticas básicas
        total_cuyes = Cuy.query.count()
        cuyes_disponibles = Cuy.query.filter_by(estado='sano').count()
        total_pozas = Poza.query.count()
        total_ventas = Venta.query.count()
        
        # Estadísticas específicas por rol
        if current_user.is_admin() or current_user.is_empleado():
            ventas_pendientes = Venta.query.filter_by(estado='pendiente').count()
            cuyes_en_tratamiento = Cuy.query.filter_by(estado='en_tratamiento').count()
            # Solo considerar ingresos de ventas completadas usando precio
            ventas_completadas = Venta.query.filter_by(estado='completada').all()
            ingresos_totales = sum(venta.precio for venta in ventas_completadas if venta.precio)
            
            admin_stats = f"""
            <div class='row mb-4'>
                <div class='col-md-3 mb-3'>
                    <div class='card text-center bg-primary text-white'>
                        <div class='card-body'>
                            <i class='fas fa-paw fa-2x mb-2'></i>
                            <h3>{total_cuyes}</h3>
                            <p>Total Cuyes</p>
                        </div>
                    </div>
                </div>
                <div class='col-md-3 mb-3'>
                    <div class='card text-center bg-success text-white'>
                        <div class='card-body'>
                            <i class='fas fa-check-circle fa-2x mb-2'></i>
                            <h3>{cuyes_disponibles}</h3>
                            <p>Disponibles</p>
                        </div>
                    </div>
                </div>
                <div class='col-md-3 mb-3'>
                    <div class='card text-center bg-warning text-white'>
                        <div class='card-body'>
                            <i class='fas fa-shopping-cart fa-2x mb-2'></i>
                            <h3>{ventas_pendientes}</h3>
                            <p>Ventas Pendientes</p>
                        </div>
                    </div>
                </div>
                <div class='col-md-3 mb-3'>
                    <div class='card text-center bg-info text-white'>
                        <div class='card-body'>
                            <i class='fas fa-coins fa-2x mb-2'></i>
                            <h4>S/ {ingresos_totales:.2f}</h4>
                            <p>Ingresos</p>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class='row mb-4'>
                <div class='col-md-3 mb-3'>
                    <div class='card text-center bg-secondary text-white'>
                        <div class='card-body'>
                            <i class='fas fa-home fa-2x mb-2'></i>
                            <h3>{total_pozas}</h3>
                            <p>Total Pozas</p>
                        </div>
                    </div>
                </div>
                <div class='col-md-3 mb-3'>
                    <div class='card text-center bg-danger text-white'>
                        <div class='card-body'>
                            <i class='fas fa-medkit fa-2x mb-2'></i>
                            <h3>{cuyes_en_tratamiento}</h3>
                            <p>En Tratamiento</p>
                        </div>
                    </div>
                </div>
                <div class='col-md-3 mb-3'>
                    <div class='card text-center bg-dark text-white'>
                        <div class='card-body'>
                            <i class='fas fa-chart-line fa-2x mb-2'></i>
                            <h3>{total_ventas}</h3>
                            <p>Total Ventas</p>
                        </div>
                    </div>
                </div>
                <div class='col-md-3 mb-3'>
                    <div class='card text-center bg-dark text-white shadow'>
                        <div class='card-body'>
                            <i class='fas fa-users fa-2x mb-2'></i>
                            <h3>{Usuario.query.filter_by(rol='cliente').count()}</h3>
                            <p>Clientes</p>
                        </div>
                    </div>
                </div>
            </div>
            """
        else:
            # Estadísticas para clientes
            mis_compras = Venta.query.filter_by(cliente_id=current_user.id).count()
            mis_compras_pendientes = Venta.query.filter_by(cliente_id=current_user.id, estado='pendiente').count()
            
            admin_stats = f"""
            <div class='row mb-4'>
                <div class='col-md-4 mb-3'>
                    <div class='card text-center bg-primary text-white'>
                        <div class='card-body'>
                            <i class='fas fa-paw fa-2x mb-2'></i>
                            <h3>{cuyes_disponibles}</h3>
                            <p>Cuyes Disponibles</p>
                        </div>
                    </div>
                </div>
                <div class='col-md-4 mb-3'>
                    <div class='card text-center bg-success text-white'>
                        <div class='card-body'>
                            <i class='fas fa-shopping-bag fa-2x mb-2'></i>
                            <h3>{mis_compras}</h3>
                            <p>Mis Compras</p>
                        </div>
                    </div>
                </div>
                <div class='col-md-4 mb-3'>
                    <div class='card text-center bg-warning text-white'>
                        <div class='card-body'>
                            <i class='fas fa-clock fa-2x mb-2'></i>
                            <h3>{mis_compras_pendientes}</h3>
                            <p>Pendientes</p>
                        </div>
                    </div>
                </div>
            </div>
            """
        
        content = f"""
        <h2>Dashboard - Sistema INIA Cuyes</h2>
        <div class='alert alert-info'>
            <i class='fas fa-user-circle'></i>
            <strong>Bienvenido, {current_user.nombre} {current_user.apellido}</strong><br>
            Rol: {current_user.rol.title()} | Último acceso: {datetime.now().strftime('%d/%m/%Y %H:%M')}
        </div>
        
        {admin_stats}
        
        <div class='row'>
            <div class='col-md-4 mb-3'>
                <div class='card'>
                    <div class='card-header bg-primary text-white'>
                        <h5><i class='fas fa-paw'></i> Gestión de Cuyes</h5>
                    </div>
                    <div class='card-body'>
                        <p>Administra el inventario de cuyes del centro.</p>
                        <div class='d-grid gap-2'>
                            <a href='/cuyes' class='btn btn-primary'>Ver Cuyes</a>"""
        
        # Solo mostrar "Agregar Cuy" para admin y empleado, no para cliente
        if current_user.is_admin() or current_user.is_empleado():
            content += """
                            <a href='/cuyes/nuevo' class='btn btn-outline-primary'>Agregar Cuy</a>"""
        
        content += """
                        </div>
                    </div>
                </div>
            </div>
            
            <div class='col-md-4 mb-3'>
                <div class='card'>
                    <div class='card-header bg-success text-white'>
                        <h5><i class='fas fa-home'></i> Gestión de Pozas</h5>
                    </div>
                    <div class='card-body'>
                        <p>Administra las pozas y su ocupación.</p>
                        <div class='d-grid gap-2'>
                            <a href='/pozas' class='btn btn-success'>Ver Pozas</a>"""
        
        # Solo mostrar "Agregar Poza" para admin y empleado, no para cliente
        if current_user.is_admin() or current_user.is_empleado():
            content += """
                            <a href='/pozas/nueva' class='btn btn-outline-success'>Agregar Poza</a>"""
        
        content += """
                        </div>
                    </div>
                </div>
            </div>
            
            <div class='col-md-4 mb-3'>
                <div class='card'>
                    <div class='card-header bg-info text-white'>
                        <h5><i class='fas fa-shopping-cart'></i> Sistema de Ventas</h5>
                    </div>
                    <div class='card-body'>
                        <p>Gestiona las ventas y pedidos de clientes.</p>
                        <div class='d-grid gap-2'>"""
        
        if current_user.is_admin() or current_user.is_empleado():
            content += """
                            <a href='/admin/ventas' class='btn btn-info'>Gestionar Ventas</a>
                            <a href='/catalogo' class='btn btn-outline-info'>Ver Catálogo</a>"""
        else:
            content += """
                            <a href='/catalogo' class='btn btn-info'>Ver Catálogo</a>
                            <a href='/mis-compras' class='btn btn-outline-info'>Mis Compras</a>"""
        
        content += """
                        </div>
                    </div>
                </div>
            </div>
        </div>
        
        <div class='row'>"""
        
        if current_user.is_admin() or current_user.is_empleado():
            content += """
            <div class='col-md-4 mb-3'>
                <div class='card'>
                    <div class='card-header bg-warning text-white'>
                        <h5><i class='fas fa-heartbeat'></i> Control Sanitario</h5>
                    </div>
                    <div class='card-body'>
                        <p>Monitoreo de salud y tratamientos.</p>
                        <div class='d-grid gap-2'>
                            <a href='/controles' class='btn btn-warning'>Ver Controles</a>
                            <a href='/tratamientos' class='btn btn-outline-warning'>Tratamientos</a>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class='col-md-4 mb-3'>
                <div class='card'>
                    <div class='card-header bg-secondary text-white'>
                        <h5><i class='fas fa-chart-bar'></i> Reportes</h5>
                    </div>
                    <div class='card-body'>
                        <p>Genera reportes y estadísticas.</p>
                        <div class='d-grid gap-2'>
                            <a href='/reportes' class='btn btn-secondary'>Ver Reportes</a>
                            <a href='/reportes/estadisticas' class='btn btn-outline-secondary'>Estadísticas</a>
                        </div>
                    </div>
                </div>
            </div>"""
            
            # Solo mostrar Administración para admin, no para empleado
            if current_user.is_admin():
                content += """
            
            <div class='col-md-4 mb-3'>
                <div class='card'>
                    <div class='card-header bg-dark text-white'>
                        <h5><i class='fas fa-cogs'></i> Administración</h5>
                    </div>
                    <div class='card-body'>
                        <p>Configuración del sistema.</p>
                        <div class='d-grid gap-2'>
                            <a href='/usuarios' class='btn btn-dark'>Usuarios</a>
                            <a href='/razas' class='btn btn-outline-dark'>Razas</a>
                        </div>
                    </div>
                </div>
            </div>"""
        else:
            content += """
            <div class='col-md-6 mb-3'>
                <div class='card'>
                    <div class='card-header bg-success text-white'>
                        <h5><i class='fas fa-leaf'></i> Información INIA</h5>
                    </div>
                    <div class='card-body'>
                        <p>Conoce más sobre nuestros cuyes y prácticas.</p>
                        <div class='d-grid gap-2'>
                            <a href='/catalogo' class='btn btn-success'>Ver Catálogo</a>
                            <a href='#' class='btn btn-outline-success'>Información Técnica</a>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class='col-md-6 mb-3'>
                <div class='card'>
                    <div class='card-header bg-primary text-white'>
                        <h5><i class='fas fa-user'></i> Mi Cuenta</h5>
                    </div>
                    <div class='card-body'>
                        <p>Gestiona tu perfil y compras.</p>
                        <div class='d-grid gap-2'>
                            <a href='/mis-compras' class='btn btn-primary'>Mis Compras</a>
                            <a href='/perfil' class='btn btn-outline-primary'>Editar Perfil</a>
                        </div>
                    </div>
                </div>
            </div>"""
        
        content += """
        </div>
        
        <div class='mt-4'>
            <div class='alert alert-light border'>
                <h6><i class='fas fa-info-circle text-primary'></i> Sistema de Gestión de Cuyes - INIA Andahuaylas</h6>
                <p class='mb-0'>Plataforma integral para la administración, venta y seguimiento de cuyes de alta calidad genética.</p>
            </div>
        </div>
        """
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    @app.route('/cuyes')
    @login_required
    def cuyes():
        cuyes_list = Cuy.query.all()
        return render_template_string(CUYES_TEMPLATE, cuyes=cuyes_list)
    
    @app.route('/cuyes/nuevo', methods=['GET', 'POST'])
    @login_required
    def nuevo_cuy():
        # Solo admin y empleados pueden registrar cuyes
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para registrar cuyes', 'danger')
            return redirect(url_for('dashboard'))
            
        form = CuyForm()
        form.raza_id.choices = [(r.id, r.nombre) for r in Raza.query.all()]
        form.poza_id.choices = [(p.id, f"{p.codigo} - {p.tipo}") for p in Poza.query.all()]
        
        if form.validate_on_submit():
            # Verificar capacidad de la poza
            poza = Poza.query.get(form.poza_id.data)
            cuyes_actuales = Cuy.query.filter_by(poza_id=poza.id, estado='sano').count() if poza else 0
            
            if poza and cuyes_actuales >= poza.capacidad:
                flash(f'La poza {poza.codigo} ha alcanzado su capacidad máxima ({poza.capacidad} cuyes). Actualmente tiene {cuyes_actuales} cuyes.', 'danger')
                return render_template_string(FORM_TEMPLATE, form=form, title="Registrar Nuevo Cuy", action=url_for('nuevo_cuy'))
            
            cuy = Cuy(
                codigo=form.codigo.data,
                sexo=form.sexo.data,
                peso_actual=form.peso_actual.data,
                precio_venta=form.precio_venta.data,
                fecha_nacimiento=form.fecha_nacimiento.data,
                raza_id=form.raza_id.data,
                poza_id=form.poza_id.data,
                observaciones=form.observaciones.data
            )
            
            db.session.add(cuy)
            db.session.commit()
            
            # Actualizar ocupación de pozas
            actualizar_ocupacion_pozas()
            
            flash('Cuy registrado exitosamente', 'success')
            return redirect(url_for('cuyes'))
        
        return render_template_string(FORM_TEMPLATE, form=form, title="Registrar Nuevo Cuy", action=url_for('nuevo_cuy'))
    
    @app.route('/razas')
    @login_required
    def razas():
        razas_list = Raza.query.all()
        content = "<h2>Gestión de Razas</h2>"
        content += f"<a href='/razas/nueva' class='btn btn-success mb-3'>Nueva Raza</a>"
        content += "<div class='row'>"
        for raza in razas_list:
            content += f"""
            <div class='col-md-4 mb-3'>
                <div class='card'>
                    <div class='card-body'>
                        <h5>{raza.nombre}</h5>
                        <p>{raza.descripcion or 'Sin descripción'}</p>
                    </div>
                </div>
            </div>
            """
        content += "</div><a href='/dashboard' class='btn btn-secondary'>Volver</a>"
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    @app.route('/razas/nueva', methods=['GET', 'POST'])
    @login_required
    def nueva_raza():
        # Solo admin puede registrar razas (más restrictivo)
        if not current_user.is_admin():
            flash('No tienes permisos para registrar razas', 'danger')
            return redirect(url_for('dashboard'))
            
        form = RazaForm()
        if form.validate_on_submit():
            raza = Raza(nombre=form.nombre.data, descripcion=form.descripcion.data)
            db.session.add(raza)
            db.session.commit()
            flash('Raza registrada exitosamente', 'success')
            return redirect(url_for('razas'))
        
        return render_template_string(FORM_TEMPLATE, form=form, title="Registrar Nueva Raza", action=url_for('nueva_raza'))
    
    @app.route('/pozas')
    @login_required
    def pozas():
        pozas_list = Poza.query.all()
        content = "<h2>Gestión de Pozas</h2>"
        content += f"<a href='/pozas/nueva' class='btn btn-warning mb-3'>Nueva Poza</a>"
        content += "<div class='row'>"
        for poza in pozas_list:
            content += f"""
            <div class='col-md-4 mb-3'>
                <div class='card'>
                    <div class='card-body'>
                        <h5>{poza.codigo}</h5>
                        <p><strong>Tipo:</strong> {poza.tipo.title()}<br>
                        <strong>Capacidad:</strong> {poza.capacidad}</p>
                    </div>
                </div>
            </div>
            """
        content += "</div><a href='/dashboard' class='btn btn-secondary'>Volver</a>"
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    @app.route('/pozas/nueva', methods=['GET', 'POST'])
    @login_required
    def nueva_poza():
        # Solo admin y empleados pueden registrar pozas
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para registrar pozas', 'danger')
            return redirect(url_for('dashboard'))
            
        form = PozaForm()
        if form.validate_on_submit():
            poza = Poza(
                codigo=form.codigo.data,
                tipo=form.tipo.data,
                capacidad=int(form.capacidad.data)
            )
            db.session.add(poza)
            db.session.commit()
            flash('Poza registrada exitosamente', 'success')
            return redirect(url_for('pozas'))
        
        return render_template_string(FORM_TEMPLATE, form=form, title="Registrar Nueva Poza", action=url_for('nueva_poza'))
    
    @app.route('/ventas')
    @login_required
    def ventas():
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para acceder a esta sección', 'danger')
            return redirect(url_for('dashboard'))
        
        # Lista de ventas para empleados/admin
        page = request.args.get('page', 1, type=int)
        ventas_list = Venta.query.order_by(Venta.fecha_registro.desc()).paginate(
            page=page, per_page=10, error_out=False
        )
        
        content = f"""
        <h2>Sistema de Ventas</h2>
        <div class='row mb-3'>
            <div class='col-md-6'>
                <a href='/ventas/nueva' class='btn btn-success'>
                    <i class='fas fa-plus'></i> Nueva Venta
                </a>
            </div>
            <div class='col-md-6 text-end'>
                <a href='/catalogo' class='btn btn-info'>
                    <i class='fas fa-store'></i> Ver Catálogo
                </a>
            </div>
        </div>
        
        <div class='card'>
            <div class='card-header'>
                <h5>Historial de Ventas</h5>
            </div>
            <div class='card-body'>
                <div class='table-responsive'>
                    <table class='table table-striped'>
                        <thead>
                            <tr>
                                <th>ID</th>
                                <th>Fecha</th>
                                <th>Cliente</th>
                                <th>Cuy</th>
                                <th>Precio</th>
                                <th>Estado</th>
                                <th>Acciones</th>
                            </tr>
                        </thead>
                        <tbody>
        """
        
        for venta in ventas_list.items:
            content += f"""
                            <tr>
                                <td>#{venta.id}</td>
                                <td>{venta.fecha_venta.strftime('%d/%m/%Y') if venta.fecha_venta else 'N/A'}</td>
                                <td>{venta.cliente.nombre_completo if venta.cliente else 'N/A'}</td>
                                <td>{venta.cuy.codigo if venta.cuy else 'N/A'}</td>
                                <td>S/ {venta.precio:.2f}</td>
                                <td>
                                    <span class='badge bg-{'success' if venta.estado == 'completada' else 'warning' if venta.estado == 'pendiente' else 'danger'}'>
                                        {venta.estado.title()}
                                    </span>
                                </td>
                                <td>
                                    <div class='btn-group' role='group'>
                                        <a href='/ventas/{venta.id}' class='btn btn-sm btn-primary'>Ver</a>
                                        {f'<form method="POST" action="/venta/{venta.id}/completar" style="display:inline;"><button type="submit" class="btn btn-sm btn-success" onclick="return confirm(\'¿Marcar como completada?\')">Completar</button></form>' if venta.estado == 'pendiente' else ''}
                                        {f'<form method="POST" action="/venta/{venta.id}/cancelar" style="display:inline;"><button type="submit" class="btn btn-sm btn-danger" onclick="return confirm(\'¿Cancelar venta?\')">Cancelar</button></form>' if venta.estado == 'pendiente' else ''}
                                    </div>
                                </td>
                            </tr>
            """
        
        content += """
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
        
        <div class='mt-3'>
            <a href='/dashboard' class='btn btn-secondary'>Volver al Dashboard</a>
        </div>
        """
        
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    @app.route('/venta/<int:venta_id>/completar', methods=['POST'])
    @login_required
    def completar_venta(venta_id):
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para esta acción', 'danger')
            return redirect(url_for('dashboard'))
        
        venta = Venta.query.get_or_404(venta_id)
        venta.estado = 'completada'
        db.session.commit()
        flash('Venta marcada como completada exitosamente', 'success')
        return redirect(url_for('ventas'))
    
    @app.route('/venta/<int:venta_id>/cancelar', methods=['POST'])
    @login_required
    def cancelar_venta(venta_id):
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para esta acción', 'danger')
            return redirect(url_for('dashboard'))
        
        venta = Venta.query.get_or_404(venta_id)
        if venta.estado == 'completada':
            flash('No se puede cancelar una venta ya completada', 'warning')
            return redirect(url_for('ventas'))
        
        venta.estado = 'cancelada'
        # Si había un cuy asignado, devolverlo al inventario
        if venta.cuy:
            venta.cuy.estado = 'sano'
        db.session.commit()
        flash('Venta cancelada exitosamente', 'success')
        return redirect(url_for('ventas'))
    
    @app.route('/catalogo')
    def catalogo():
        # Catálogo público de cuyes disponibles para venta
        cuyes_disponibles = Cuy.query.filter_by(estado='sano').filter(Cuy.peso_actual >= 0.8).all()
        
        content = """
        <h2><i class='fas fa-store'></i> Catálogo de Cuyes INIA</h2>
        <p class='text-muted mb-4'>Cuyes de calidad disponibles para venta</p>
        
        <div class='row'>
        """
        
        for cuy in cuyes_disponibles:
            precio = getattr(cuy, 'precio_venta', None) or (cuy.peso_actual * 25.0 if cuy.peso_actual else 50.0)
            content += f"""
            <div class='col-md-4 mb-4'>
                <div class='card h-100'>
                    <div class='card-header'>
                        <h5 class='mb-0'>Cuy {cuy.codigo}</h5>
                    </div>
                    <div class='card-body'>
                        <p><strong>Raza:</strong> {cuy.raza_obj.nombre if cuy.raza_obj else 'N/A'}</p>
                        <p><strong>Sexo:</strong> {cuy.sexo.title()}</p>
                        <p><strong>Peso:</strong> {cuy.peso_actual:.2f} kg</p>
                        <p><strong>Estado:</strong> <span class='badge bg-success'>{cuy.estado.title()}</span></p>
                        <h4 class='text-success'>S/ {precio:.2f}</h4>
                    </div>
                    <div class='card-footer'>
                        <a href='/comprar/{cuy.id}' class='btn btn-success w-100'>
                            <i class='fas fa-shopping-cart'></i> Comprar
                        </a>
                    </div>
                </div>
            </div>
            """
        
        if not cuyes_disponibles:
            content += """
            <div class='col-12'>
                <div class='alert alert-info text-center'>
                    <i class='fas fa-info-circle fa-2x mb-3'></i>
                    <h4>No hay cuyes disponibles</h4>
                    <p>En este momento no tenemos cuyes disponibles para venta. Por favor, vuelve más tarde.</p>
                </div>
            </div>
            """
        
        content += """
        </div>
        
        <div class='mt-4 text-center'>
            <a href='/' class='btn btn-secondary'>Volver al Inicio</a>
        </div>
        """
        
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    @app.route('/comprar/<int:cuy_id>', methods=['GET', 'POST'])
    @login_required
    def comprar_cuy(cuy_id):
        if not current_user.is_cliente():
            flash('Solo los clientes pueden realizar compras', 'danger')
            return redirect(url_for('catalogo'))
        
        cuy = Cuy.query.get_or_404(cuy_id)
        
        if cuy.estado != 'sano':
            flash('Este cuy ya no está disponible', 'danger')
            return redirect(url_for('catalogo'))
        
        if request.method == 'POST':
            precio = getattr(cuy, 'precio_venta', None) or (cuy.peso_actual * 25.0 if cuy.peso_actual else 50.0)
            
            venta = Venta(
                cliente_id=current_user.id,
                cuy_id=cuy_id,
                fecha_venta=datetime.now().date(),
                precio=precio,
                peso_venta=cuy.peso_actual,
                metodo_pago=request.form.get('metodo_pago', 'transferencia'),
                estado='pendiente'
            )
            
            # Cambiar estado del cuy
            cuy.estado = 'vendido'
            
            db.session.add(venta)
            db.session.commit()
            
            flash('¡Compra realizada exitosamente! Te contactaremos pronto.', 'success')
            return redirect(url_for('mis_compras'))
        
        precio = getattr(cuy, 'precio_venta', None) or (cuy.peso_actual * 25.0 if cuy.peso_actual else 50.0)
        
        content = f"""
        <h2>Confirmar Compra</h2>
        <div class='row'>
            <div class='col-md-6'>
                <div class='card'>
                    <div class='card-header'>
                        <h5>Detalles del Cuy</h5>
                    </div>
                    <div class='card-body'>
                        <p><strong>Código:</strong> {cuy.codigo}</p>
                        <p><strong>Raza:</strong> {cuy.raza.nombre if cuy.raza else 'N/A'}</p>
                        <p><strong>Sexo:</strong> {cuy.sexo.title()}</p>
                        <p><strong>Peso:</strong> {cuy.peso_actual:.2f} kg</p>
                        <h4 class='text-success'>Precio: S/ {precio:.2f}</h4>
                    </div>
                </div>
            </div>
            <div class='col-md-6'>
                <div class='card'>
                    <div class='card-header'>
                        <h5>Datos de Compra</h5>
                    </div>
                    <div class='card-body'>
                        <form method='POST'>
                            <div class='mb-3'>
                                <label class='form-label'>Cliente</label>
                                <input type='text' class='form-control' value='{current_user.nombre_completo}' readonly>
                            </div>
                            <div class='mb-3'>
                                <label class='form-label'>Email</label>
                                <input type='email' class='form-control' value='{current_user.email}' readonly>
                            </div>
                            <div class='mb-3'>
                                <label class='form-label'>Método de Pago</label>
                                <select name='metodo_pago' class='form-select' required>
                                    <option value='transferencia'>Transferencia Bancaria</option>
                                    <option value='efectivo'>Efectivo</option>
                                    <option value='tarjeta'>Tarjeta</option>
                                </select>
                            </div>
                            <button type='submit' class='btn btn-success w-100'>
                                <i class='fas fa-check'></i> Confirmar Compra
                            </button>
                        </form>
                    </div>
                </div>
            </div>
        </div>
        
        <div class='mt-3'>
            <a href='/catalogo' class='btn btn-secondary'>Volver al Catálogo</a>
        </div>
        """
        
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    @app.route('/mis-compras')
    @login_required
    def mis_compras():
        if not current_user.is_cliente():
            flash('Acceso denegado', 'danger')
            return redirect(url_for('dashboard'))
        
        compras = Venta.query.filter_by(cliente_id=current_user.id).order_by(Venta.fecha_registro.desc()).all()
        
        content = """
        <h2>Mis Compras</h2>
        
        <div class='card'>
            <div class='card-body'>
                <div class='table-responsive'>
                    <table class='table'>
                        <thead>
                            <tr>
                                <th>Fecha</th>
                                <th>Cuy</th>
                                <th>Precio</th>
                                <th>Estado</th>
                                <th>Acciones</th>
                            </tr>
                        </thead>
                        <tbody>
        """
        
        for compra in compras:
            content += f"""
                            <tr>
                                <td>{compra.fecha_venta.strftime('%d/%m/%Y') if compra.fecha_venta else 'N/A'}</td>
                                <td>{compra.cuy.codigo if compra.cuy else 'N/A'}</td>
                                <td>S/ {compra.precio:.2f}</td>
                                <td>
                                    <span class='badge bg-{'success' if compra.estado == 'completada' else 'warning' if compra.estado == 'pendiente' else 'danger'}'>
                                        {compra.estado.title()}
                                    </span>
                                </td>
                                <td>
                                    <a href='/compra/{compra.id}' class='btn btn-sm btn-primary'>Ver Detalle</a>
                                </td>
                            </tr>
            """
        
        if not compras:
            content += """
                            <tr>
                                <td colspan='5' class='text-center'>
                                    <div class='alert alert-info'>
                                        No has realizado compras aún.
                                        <a href='/catalogo' class='btn btn-sm btn-primary ms-2'>Ver Catálogo</a>
                                    </div>
                                </td>
                            </tr>
            """
        
        content += """
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
        
        <div class='mt-3'>
            <a href='/catalogo' class='btn btn-success me-2'>Ver Catálogo</a>
            <a href='/dashboard' class='btn btn-secondary'>Volver al Dashboard</a>
        </div>
        """
        
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    @app.route('/compra/<int:compra_id>')
    @login_required
    def detalle_compra(compra_id):
        if not current_user.is_cliente():
            flash('Acceso denegado', 'danger')
            return redirect(url_for('dashboard'))
        
        compra = Venta.query.filter_by(id=compra_id, cliente_id=current_user.id).first_or_404()
        
        content = f"""
        <div class='row justify-content-center'>
            <div class='col-md-8'>
                <div class='card'>
                    <div class='card-header bg-primary text-white'>
                        <h4><i class='fas fa-receipt'></i> Detalle de Compra #{compra.id}</h4>
                    </div>
                    <div class='card-body'>
                        <div class='row'>
                            <div class='col-md-6'>
                                <h5>Información del Cuy</h5>
                                <table class='table table-sm'>
                                    <tr>
                                        <td><strong>Código:</strong></td>
                                        <td>{compra.cuy.codigo if compra.cuy else 'N/A'}</td>
                                    </tr>
                                    <tr>
                                        <td><strong>Raza:</strong></td>
                                        <td>{compra.cuy.raza_obj.nombre if compra.cuy and compra.cuy.raza_obj else 'N/A'}</td>
                                    </tr>
                                    <tr>
                                        <td><strong>Sexo:</strong></td>
                                        <td>{compra.cuy.sexo.title() if compra.cuy else 'N/A'}</td>
                                    </tr>
                                    <tr>
                                        <td><strong>Peso en venta:</strong></td>
                                        <td>{compra.peso_venta if compra.peso_venta else compra.cuy.peso_actual if compra.cuy else 'N/A'} kg</td>
                                    </tr>
                                </table>
                            </div>
                            <div class='col-md-6'>
                                <h5>Información de la Compra</h5>
                                <table class='table table-sm'>
                                    <tr>
                                        <td><strong>Fecha de compra:</strong></td>
                                        <td>{compra.fecha_venta.strftime('%d/%m/%Y') if compra.fecha_venta else 'N/A'}</td>
                                    </tr>
                                    <tr>
                                        <td><strong>Precio:</strong></td>
                                        <td><h5 class='text-success'>S/ {compra.precio:.2f}</h5></td>
                                    </tr>
                                    <tr>
                                        <td><strong>Método de pago:</strong></td>
                                        <td>{compra.metodo_pago.title()}</td>
                                    </tr>
                                    <tr>
                                        <td><strong>Estado:</strong></td>
                                        <td>
                                            <span class='badge bg-{"success" if compra.estado == "completada" else "warning" if compra.estado == "pendiente" else "danger"} fs-6'>
                                                {compra.estado.title()}
                                            </span>
                                        </td>
                                    </tr>
                                </table>
                            </div>
                        </div>
                        
                        {f"<div class='mt-3'><h5>Observaciones:</h5><p class='text-muted'>{compra.observaciones}</p></div>" if compra.observaciones else ""}
                        
                        <div class='mt-4 text-center'>
                            {f'<div class="alert alert-success"><i class="fas fa-check-circle"></i> Compra completada exitosamente</div>' if compra.estado == 'completada' else ''}
                            {f'<div class="alert alert-warning"><i class="fas fa-clock"></i> Tu compra está siendo procesada. Te contactaremos pronto.</div>' if compra.estado == 'pendiente' else ''}
                            {f'<div class="alert alert-danger"><i class="fas fa-times-circle"></i> Compra cancelada.</div>' if compra.estado == 'cancelada' else ''}
                        </div>
                        
                        <div class='d-grid gap-2 d-md-flex justify-content-md-center mt-4'>
                            <a href='/mis-compras' class='btn btn-secondary me-md-2'>
                                <i class='fas fa-arrow-left'></i> Volver a Mis Compras
                            </a>
                            <a href='/catalogo' class='btn btn-primary me-md-2'>
                                <i class='fas fa-shopping-cart'></i> Seguir Comprando
                            </a>
                            {f'<a href="/factura/{compra.id}" class="btn btn-success"><i class="fas fa-download"></i> Descargar Factura</a>' if compra.estado == 'completada' else ''}
                        </div>
                    </div>
                </div>
            </div>
        </div>
        """
        
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    @app.route('/factura/<int:compra_id>')
    @login_required
    def descargar_factura(compra_id):
        if not current_user.is_cliente():
            flash('Acceso denegado', 'danger')
            return redirect(url_for('dashboard'))
        
        # Verificar que la compra pertenece al cliente actual y está completada
        compra = Venta.query.filter_by(id=compra_id, cliente_id=current_user.id).first_or_404()
        
        if compra.estado != 'completada':
            flash('Solo puedes descargar facturas de compras completadas', 'warning')
            return redirect(url_for('detalle_compra', compra_id=compra_id))
        
        # Crear el documento PDF para la factura
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.8*inch, bottomMargin=0.8*inch)
        
        # Estilos
        styles = getSampleStyleSheet()
        
        # Estilo para el título de la factura
        title_style = ParagraphStyle(
            'FacturaTitle',
            parent=styles['Title'],
            fontSize=22,
            textColor=colors.darkgreen,
            spaceAfter=20,
            alignment=1  # Center
        )
        
        # Estilo para información de la empresa
        empresa_style = ParagraphStyle(
            'EmpresaStyle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.darkblue,
            spaceAfter=15,
            alignment=1  # Center
        )
        
        # Estilo para datos de la factura
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=10
        )
        
        # Estilo para totales
        total_style = ParagraphStyle(
            'TotalStyle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.darkgreen,
            spaceAfter=10
        )
        
        # Contenido del documento
        story = []
        
        # Encabezado de la empresa
        empresa_info = """
        <b>INSTITUTO NACIONAL DE INNOVACIÓN AGRARIA</b><br/>
        <b>INIA - ANDAHUAYLAS</b><br/>
        Sistema de Gestión de Cuyes<br/>
        RUC: 20123456789<br/>
        Teléfono: (083) 123-456<br/>
        Email: cuyes@inia.gob.pe
        """
        empresa_para = Paragraph(empresa_info, empresa_style)
        story.append(empresa_para)
        story.append(Spacer(1, 20))
        
        # Título de la factura
        title = Paragraph(f"FACTURA ELECTRÓNICA N° {compra.id:06d}", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Información del cliente y factura
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
        fecha_venta = compra.fecha_venta.strftime('%d/%m/%Y') if compra.fecha_venta else 'N/A'
        
        # Crear tabla con información del cliente y factura
        client_data = [
            ['DATOS DEL CLIENTE', 'DATOS DE LA FACTURA'],
            [f'Nombre: {current_user.nombre} {current_user.apellido}', f'Fecha de Emisión: {fecha_actual}'],
            [f'Email: {current_user.email}', f'Fecha de Venta: {fecha_venta}'],
            [f'Teléfono: {current_user.telefono or "No registrado"}', f'Método de Pago: {compra.metodo_pago.title()}'],
            [f'Dirección: {current_user.direccion or "No registrada"}', f'Estado: {compra.estado.title()}']
        ]
        
        client_table = Table(client_data, colWidths=[3*inch, 3*inch])
        client_table.setStyle(TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Contenido
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(client_table)
        story.append(Spacer(1, 30))
        
        # Detalles del producto
        detalle_title = Paragraph("<b>DETALLE DE LA COMPRA</b>", info_style)
        story.append(detalle_title)
        story.append(Spacer(1, 10))
        
        # Tabla de productos
        product_data = [
            ['Descripción', 'Código', 'Raza', 'Sexo', 'Peso (kg)', 'Precio Unit.', 'Cantidad', 'Total']
        ]
        
        # Datos del cuy
        cuy_descripcion = f"Cuy {compra.cuy.codigo}" if compra.cuy else "Cuy N/A"
        cuy_codigo = compra.cuy.codigo if compra.cuy else "N/A"
        cuy_raza = compra.cuy.raza_obj.nombre if compra.cuy and compra.cuy.raza_obj else "N/A"
        cuy_sexo = compra.cuy.sexo.title() if compra.cuy else "N/A"
        cuy_peso = f"{compra.peso_venta:.2f}" if compra.peso_venta else (f"{compra.cuy.peso_actual:.2f}" if compra.cuy and compra.cuy.peso_actual else "N/A")
        precio_unitario = f"S/ {compra.precio:.2f}"
        cantidad = "1"
        total_item = f"S/ {compra.precio:.2f}"
        
        product_data.append([
            cuy_descripcion,
            cuy_codigo,
            cuy_raza,
            cuy_sexo,
            cuy_peso,
            precio_unitario,
            cantidad,
            total_item
        ])
        
        product_table = Table(product_data, repeatRows=1)
        product_table.setStyle(TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Contenido
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(product_table)
        story.append(Spacer(1, 30))
        
        # Resumen de totales
        total_data = [
            ['', 'SUBTOTAL:', f'S/ {compra.precio:.2f}'],
            ['', 'IGV (18%):', f'S/ {compra.precio * 0.18:.2f}'],
            ['', 'TOTAL A PAGAR:', f'S/ {compra.precio * 1.18:.2f}']
        ]
        
        total_table = Table(total_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        total_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (1, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 0), (-1, -1), 12),
            ('TEXTCOLOR', (1, 2), (-1, 2), colors.darkgreen),
            ('FONTSIZE', (1, 2), (-1, 2), 14),
            ('LINEABOVE', (1, 2), (-1, 2), 2, colors.darkgreen),
            ('TOPPADDING', (1, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (1, 0), (-1, -1), 8),
        ]))
        
        story.append(total_table)
        story.append(Spacer(1, 40))
        
        # Nota final
        nota_text = """
        <b>NOTA IMPORTANTE:</b><br/>
        Esta factura corresponde a la compra de cuy de calidad genética del INIA Andahuaylas.
        El producto ha sido criado bajo estrictos controles sanitarios y de alimentación.
        Para cualquier consulta, comuníquese con nosotros a través de los medios de contacto indicados.
        """
        nota_para = Paragraph(nota_text, info_style)
        story.append(nota_para)
        
        # Pie de página
        story.append(Spacer(1, 20))
        pie_text = f"<i>Factura generada electrónicamente el {fecha_actual} - Sistema INIA Cuyes v1.0</i>"
        pie_para = Paragraph(pie_text, ParagraphStyle('Pie', parent=styles['Normal'], fontSize=8, alignment=1, textColor=colors.grey))
        story.append(pie_para)
        
        # Generar el PDF
        doc.build(story)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'factura_{compra.id:06d}_{datetime.now().strftime("%Y%m%d")}.pdf',
            mimetype='application/pdf'
        )
    
    @app.route('/reportes')
    @login_required
    def reportes():
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para acceder a reportes', 'danger')
            return redirect(url_for('dashboard'))
        
        content = """
        <h2>Reportes y Estadísticas</h2>
        
        <div class='row mb-4'>
            <div class='col-12'>
                <div class='alert alert-info'>
                    <i class='fas fa-info-circle'></i>
                    <strong>Sistema de Reportes INIA</strong><br>
                    Genera reportes detallados del inventario, ventas y controles sanitarios
                </div>
            </div>
        </div>
        
        <div class='row'>
            <div class='col-md-6 mb-4'>
                <div class='card'>
                    <div class='card-header bg-primary text-white'>
                        <h5><i class='fas fa-users'></i> Reporte de Cuyes</h5>
                    </div>
                    <div class='card-body'>
                        <p>Inventario completo de cuyes con detalles de raza, peso y estado.</p>
                        <div class='d-grid gap-2'>
                            <a href='/reportes/cuyes/pdf' class='btn btn-danger'>
                                <i class='fas fa-file-pdf'></i> Generar PDF
                            </a>
                            <a href='/reportes/cuyes/excel' class='btn btn-success'>
                                <i class='fas fa-file-excel'></i> Generar Excel
                            </a>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class='col-md-6 mb-4'>
                <div class='card'>
                    <div class='card-header bg-info text-white'>
                        <h5><i class='fas fa-shopping-cart'></i> Reporte de Ventas</h5>
                    </div>
                    <div class='card-body'>
                        <p>Historial de ventas con datos de clientes y transacciones.</p>
                        <div class='d-grid gap-2'>
                            <a href='/reportes/ventas/pdf' class='btn btn-danger'>
                                <i class='fas fa-file-pdf'></i> Generar PDF
                            </a>
                            <a href='/reportes/ventas/excel' class='btn btn-success'>
                                <i class='fas fa-file-excel'></i> Generar Excel
                            </a>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class='col-md-6 mb-4'>
                <div class='card'>
                    <div class='card-header bg-success text-white'>
                        <h5><i class='fas fa-home'></i> Reporte de Pozas</h5>
                    </div>
                    <div class='card-body'>
                        <p>Estado y ocupación de las pozas del centro de crianza.</p>
                        <div class='d-grid gap-2'>
                            <a href='/reportes/pozas/pdf' class='btn btn-danger'>
                                <i class='fas fa-file-pdf'></i> Generar PDF
                            </a>
                            <a href='/reportes/pozas/excel' class='btn btn-success'>
                                <i class='fas fa-file-excel'></i> Generar Excel
                            </a>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class='col-md-6 mb-4'>
                <div class='card'>
                    <div class='card-header bg-warning text-white'>
                        <h5><i class='fas fa-heartbeat'></i> Estadísticas Generales</h5>
                    </div>
                    <div class='card-body'>
                        <p>Resumen estadístico del sistema de gestión de cuyes.</p>
                        <div class='d-grid gap-2'>
                            <a href='/reportes/estadisticas' class='btn btn-warning text-dark'>
                                <i class='fas fa-chart-bar'></i> Ver Estadísticas
                            </a>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        
        <div class='mt-3'>
            <a href='/dashboard' class='btn btn-secondary'>Volver al Dashboard</a>
        </div>
        """
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    @app.route('/reportes/cuyes/pdf')
    @login_required
    def reporte_cuyes_pdf():
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para generar reportes', 'danger')
            return redirect(url_for('dashboard'))
        
        # Obtener datos de cuyes
        cuyes = Cuy.query.all()
        
        # Crear el documento PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            textColor=colors.darkblue,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Contenido del documento
        story = []
        
        # Título
        title = Paragraph("REPORTE DE CUYES - INIA ANDAHUAYLAS", title_style)
        story.append(title)
        
        # Información del reporte
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20
        )
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
        info_text = f"<b>Fecha del reporte:</b> {fecha_actual}<br/><b>Total de cuyes:</b> {len(cuyes)}"
        info_para = Paragraph(info_text, info_style)
        story.append(info_para)
        story.append(Spacer(1, 12))
        
        # Crear tabla de datos
        data = [['ID', 'Código', 'Raza', 'Sexo', 'Peso (kg)', 'Estado', 'Poza', 'Fecha Nac.']]
        
        for cuy in cuyes:
            poza_nombre = cuy.poza.codigo if cuy.poza else "Sin asignar"
            fecha_nac = cuy.fecha_nacimiento.strftime('%d/%m/%Y') if cuy.fecha_nacimiento else "No registrada"
            raza_nombre = cuy.raza_obj.nombre if cuy.raza_obj else (cuy.raza or "N/A")
            peso = f"{cuy.peso_actual:.2f}" if cuy.peso_actual else "N/A"
            
            data.append([
                str(cuy.id),
                cuy.codigo or "N/A",
                raza_nombre,
                cuy.sexo.title(),
                peso,
                cuy.estado.replace('_', ' ').title(),
                poza_nombre,
                fecha_nac
            ])
        
        # Crear y estilizar la tabla
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Contenido
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Alternar colores de filas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
        ]))
        
        story.append(table)
        
        # Generar el PDF
        doc.build(story)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_cuyes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf',
            mimetype='application/pdf'
        )
    
    @app.route('/reportes/cuyes/excel')
    @login_required
    def reporte_cuyes_excel():
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para generar reportes', 'danger')
            return redirect(url_for('dashboard'))
        
        # Obtener datos de cuyes
        cuyes = Cuy.query.all()
        
        # Preparar datos para el DataFrame
        data = []
        for cuy in cuyes:
            poza_nombre = cuy.poza.codigo if cuy.poza else "Sin asignar"
            fecha_nac = cuy.fecha_nacimiento.strftime('%d/%m/%Y') if cuy.fecha_nacimiento else "No registrada"
            fecha_registro = cuy.fecha_registro.strftime('%d/%m/%Y') if cuy.fecha_registro else "No registrada"
            raza_nombre = cuy.raza_obj.nombre if cuy.raza_obj else (cuy.raza or "N/A")
            peso = cuy.peso_actual if cuy.peso_actual else 0
            
            data.append({
                'ID': cuy.id,
                'Código': cuy.codigo or "N/A",
                'Raza': raza_nombre,
                'Sexo': cuy.sexo.title(),
                'Peso (kg)': peso,
                'Estado': cuy.estado.replace('_', ' ').title(),
                'Poza': poza_nombre,
                'Fecha Nacimiento': fecha_nac,
                'Fecha Registro': fecha_registro,
                'Observaciones': cuy.observaciones or "Ninguna"
            })
        
        # Crear DataFrame
        df = pd.DataFrame(data)
        
        # Crear buffer para el archivo Excel
        buffer = io.BytesIO()
        
        # Crear el archivo Excel con formato
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Escribir los datos
            df.to_excel(writer, sheet_name='Reporte_Cuyes', index=False)
            
            # Obtener el workbook y worksheet para aplicar formato
            workbook = writer.book
            worksheet = writer.sheets['Reporte_Cuyes']
            
            # Aplicar formato al encabezado
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Agregar información del reporte en una hoja separada
            info_data = {
                'Información del Reporte': [
                    'Fecha de generación',
                    'Total de cuyes',
                    'Cuyes disponibles',
                    'Cuyes en tratamiento',
                    'Generado por'
                ],
                'Valor': [
                    datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                    len(cuyes),
                    len([c for c in cuyes if c.estado == 'sano']),
                    len([c for c in cuyes if c.estado == 'en_tratamiento']),
                    f"{current_user.nombre} {current_user.apellido}"
                ]
            }
            
            info_df = pd.DataFrame(info_data)
            info_df.to_excel(writer, sheet_name='Información', index=False)
            
            # Formatear hoja de información
            info_sheet = writer.sheets['Información']
            for cell in info_sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            for column in info_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                info_sheet.column_dimensions[column_letter].width = adjusted_width
        
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_cuyes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    @app.route('/reportes/ventas/pdf')
    @login_required
    def reporte_ventas_pdf():
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para generar reportes', 'danger')
            return redirect(url_for('dashboard'))
        
        # Obtener datos de ventas
        ventas = Venta.query.order_by(Venta.fecha_venta.desc()).all()
        total_ingresos = sum(venta.precio for venta in ventas if venta.precio and venta.estado == 'completada')
        
        # Crear el documento PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            textColor=colors.darkgreen,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Contenido del documento
        story = []
        
        # Título
        title = Paragraph("REPORTE DE VENTAS - INIA ANDAHUAYLAS", title_style)
        story.append(title)
        
        # Información del reporte
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20
        )
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
        info_text = f"<b>Fecha del reporte:</b> {fecha_actual}<br/><b>Total de ventas:</b> {len(ventas)}<br/><b>Ingresos totales:</b> S/ {total_ingresos:.2f}"
        info_para = Paragraph(info_text, info_style)
        story.append(info_para)
        story.append(Spacer(1, 12))
        
        # Crear tabla de datos
        data = [['ID', 'Cliente', 'Email', 'Teléfono', 'Cuy', 'Precio', 'Estado', 'Fecha']]
        
        for venta in ventas:
            fecha_venta = venta.fecha_venta.strftime('%d/%m/%Y') if venta.fecha_venta else "No registrada"
            cuy_codigo = venta.cuy.codigo if venta.cuy else "N/A"
            
            data.append([
                str(venta.id),
                venta.nombre_cliente or "N/A",
                venta.email_cliente or "N/A",
                venta.telefono_cliente or "N/A",
                cuy_codigo,
                f"S/ {venta.precio:.2f}" if venta.precio else "N/A",
                venta.estado.title(),
                fecha_venta
            ])
        
        # Crear y estilizar la tabla
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Contenido
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Alternar colores de filas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightgreen, colors.white])
        ]))
        
        story.append(table)
        
        # Generar el PDF
        doc.build(story)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf',
            mimetype='application/pdf'
        )
    
    # Sistema de Control Sanitario
    @app.route('/controles')
    @login_required
    def controles():
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para acceder a controles sanitarios', 'danger')
            return redirect(url_for('dashboard'))
        
        controles = Control.query.order_by(Control.fecha_control.desc()).all()
        
        content = f"""
        <h2>Control Sanitario de Cuyes</h2>
        
        <div class='row mb-4'>
            <div class='col-12'>
                <div class='alert alert-success'>
                    <i class='fas fa-heartbeat'></i>
                    <strong>Sistema de Monitoreo Sanitario</strong><br>
                    Registra y supervisa el estado de salud de los cuyes
                </div>
            </div>
        </div>
        
        <div class='mb-3'>
            <a href='/controles/nuevo' class='btn btn-primary'>
                <i class='fas fa-plus'></i> Nuevo Control
            </a>
            <a href='/tratamientos' class='btn btn-warning'>
                <i class='fas fa-medkit'></i> Ver Tratamientos
            </a>
        </div>
        
        <div class='table-responsive'>
            <table class='table table-striped'>
                <thead class='table-dark'>
                    <tr>
                        <th>Fecha</th>
                        <th>Cuy</th>
                        <th>Peso (g)</th>
                        <th>Estado Salud</th>
                        <th>Observaciones</th>
                        <th>Responsable</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for control in controles:
            estado_badge = "success" if control.estado_salud == "buena" else "warning" if control.estado_salud == "regular" else "danger"
            fecha = control.fecha_control.strftime('%d/%m/%Y') if control.fecha_control else "No registrada"
            responsable = control.usuario.nombre if control.usuario else "No asignado"
            
            content += f"""
                    <tr>
                        <td>{fecha}</td>
                        <td>{control.cuy.numero if control.cuy else 'N/A'}</td>
                        <td>{control.peso}</td>
                        <td><span class='badge bg-{estado_badge}'>{control.estado_salud.title()}</span></td>
                        <td>{control.observaciones or 'Sin observaciones'}</td>
                        <td>{responsable}</td>
                    </tr>
            """
        
        if not controles:
            content += """
                    <tr>
                        <td colspan='6' class='text-center'>
                            <div class='alert alert-info'>
                                No hay controles sanitarios registrados.
                                <a href='/controles/nuevo' class='btn btn-sm btn-primary ms-2'>Crear Primer Control</a>
                            </div>
                        </td>
                    </tr>
            """
        
        content += """
                </tbody>
            </table>
        </div>
        
        <div class='mt-3'>
            <a href='/dashboard' class='btn btn-secondary'>Volver al Dashboard</a>
        </div>
        """
        
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    @app.route('/controles/nuevo', methods=['GET', 'POST'])
    @login_required
    def nuevo_control():
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para crear controles', 'danger')
            return redirect(url_for('dashboard'))
        
        form = ControlForm()
        form.cuy_id.choices = [(c.id, f"{c.numero} - {c.raza}") for c in Cuy.query.all()]
        
        if form.validate_on_submit():
            control = Control(
                cuy_id=form.cuy_id.data,
                fecha_control=form.fecha_control.data,
                peso=form.peso.data,
                estado_salud=form.estado_salud.data,
                observaciones=form.observaciones.data,
                usuario_id=current_user.id
            )
            db.session.add(control)
            
            # Actualizar peso del cuy
            cuy = Cuy.query.get(form.cuy_id.data)
            if cuy:
                cuy.peso = form.peso.data
            
            db.session.commit()
            flash('Control sanitario registrado exitosamente', 'success')
            return redirect(url_for('controles'))
        
        return render_template_string(FORM_TEMPLATE, form=form, title="Nuevo Control Sanitario", action=url_for('nuevo_control'))
    
    @app.route('/tratamientos')
    @login_required
    def tratamientos():
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para acceder a tratamientos', 'danger')
            return redirect(url_for('dashboard'))
        
        tratamientos = Tratamiento.query.order_by(Tratamiento.fecha_inicio.desc()).all()
        
        content = f"""
        <h2>Gestión de Tratamientos</h2>
        
        <div class='row mb-4'>
            <div class='col-12'>
                <div class='alert alert-warning'>
                    <i class='fas fa-medkit'></i>
                    <strong>Sistema de Tratamientos Médicos</strong><br>
                    Administra tratamientos y medicamentos para cuyes
                </div>
            </div>
        </div>
        
        <div class='mb-3'>
            <a href='/tratamientos/nuevo' class='btn btn-warning'>
                <i class='fas fa-plus'></i> Nuevo Tratamiento
            </a>
            <a href='/controles' class='btn btn-success'>
                <i class='fas fa-heartbeat'></i> Ver Controles
            </a>
        </div>
        
        <div class='table-responsive'>
            <table class='table table-striped'>
                <thead class='table-dark'>
                    <tr>
                        <th>Cuy</th>
                        <th>Medicamento</th>
                        <th>Dosis</th>
                        <th>Fecha Inicio</th>
                        <th>Fecha Fin</th>
                        <th>Estado</th>
                        <th>Responsable</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for tratamiento in tratamientos:
            estado_badge = "success" if tratamiento.estado == "completado" else "warning" if tratamiento.estado == "en_curso" else "danger"
            fecha_inicio = tratamiento.fecha_inicio.strftime('%d/%m/%Y') if tratamiento.fecha_inicio else "No registrada"
            fecha_fin = tratamiento.fecha_fin.strftime('%d/%m/%Y') if tratamiento.fecha_fin else "En curso"
            responsable = tratamiento.usuario.nombre if tratamiento.usuario else "No asignado"
            
            content += f"""
                    <tr>
                        <td>{tratamiento.cuy.numero if tratamiento.cuy else 'N/A'}</td>
                        <td>{tratamiento.medicamento}</td>
                        <td>{tratamiento.dosis}</td>
                        <td>{fecha_inicio}</td>
                        <td>{fecha_fin}</td>
                        <td><span class='badge bg-{estado_badge}'>{tratamiento.estado.replace('_', ' ').title()}</span></td>
                        <td>{responsable}</td>
                    </tr>
            """
        
        if not tratamientos:
            content += """
                    <tr>
                        <td colspan='7' class='text-center'>
                            <div class='alert alert-info'>
                                No hay tratamientos registrados.
                                <a href='/tratamientos/nuevo' class='btn btn-sm btn-warning ms-2'>Crear Primer Tratamiento</a>
                            </div>
                        </td>
                    </tr>
            """
        
        content += """
                </tbody>
            </table>
        </div>
        
        <div class='mt-3'>
            <a href='/dashboard' class='btn btn-secondary'>Volver al Dashboard</a>
        </div>
        """
        
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    @app.route('/tratamientos/nuevo', methods=['GET', 'POST'])
    @login_required
    def nuevo_tratamiento():
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para crear tratamientos', 'danger')
            return redirect(url_for('dashboard'))
        
        form = TratamientoForm()
        form.cuy_id.choices = [(c.id, f"{c.numero} - {c.raza}") for c in Cuy.query.all()]
        
        if form.validate_on_submit():
            tratamiento = Tratamiento(
                cuy_id=form.cuy_id.data,
                medicamento=form.medicamento.data,
                dosis=form.dosis.data,
                fecha_inicio=form.fecha_inicio.data,
                fecha_fin=form.fecha_fin.data,
                observaciones=form.observaciones.data,
                usuario_id=current_user.id
            )
            
            # Cambiar estado del cuy a en_tratamiento
            cuy = Cuy.query.get(form.cuy_id.data)
            if cuy:
                cuy.estado = 'en_tratamiento'
            
            db.session.add(tratamiento)
            db.session.commit()
            flash('Tratamiento registrado exitosamente', 'success')
            return redirect(url_for('tratamientos'))
        
        return render_template_string(FORM_TEMPLATE, form=form, title="Nuevo Tratamiento", action=url_for('nuevo_tratamiento'))
    
    # Reportes adicionales
    @app.route('/reportes/ventas/excel')
    @login_required
    def reporte_ventas_excel():
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para generar reportes', 'danger')
            return redirect(url_for('dashboard'))
        
        # Obtener datos de ventas
        ventas = Venta.query.order_by(Venta.fecha_venta.desc()).all()
        
        # Preparar datos para el DataFrame
        data = []
        for venta in ventas:
            fecha_venta = venta.fecha_venta.strftime('%d/%m/%Y') if venta.fecha_venta else "No registrada"
            cuy_codigo = venta.cuy.codigo if venta.cuy else "N/A"
            
            data.append({
                'ID': venta.id,
                'Cliente': venta.nombre_cliente or "N/A",
                'Email': venta.email_cliente or "N/A",
                'Teléfono': venta.telefono_cliente or "N/A",
                'Cuy (Código)': cuy_codigo,
                'Precio (S/)': venta.precio if venta.precio else 0,
                'Peso Venta (kg)': venta.peso_venta if venta.peso_venta else 0,
                'Método Pago': venta.metodo_pago.title() if venta.metodo_pago else "N/A",
                'Estado': venta.estado.title(),
                'Fecha Venta': fecha_venta,
                'Fecha Registro': venta.fecha_registro.strftime('%d/%m/%Y') if venta.fecha_registro else "N/A",
                'Vendido Por': venta.vendedor.nombre_completo if venta.vendedor else "N/A",
                'Observaciones': venta.observaciones or "Ninguna"
            })
        
        # Crear DataFrame
        df = pd.DataFrame(data)
        
        # Crear buffer para el archivo Excel
        buffer = io.BytesIO()
        
        # Crear el archivo Excel con formato
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Escribir los datos
            df.to_excel(writer, sheet_name='Reporte_Ventas', index=False)
            
            # Obtener el workbook y worksheet para aplicar formato
            workbook = writer.book
            worksheet = writer.sheets['Reporte_Ventas']
            
            # Aplicar formato al encabezado
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Agregar resumen estadístico
            total_ingresos = sum(v.precio for v in ventas if v.precio and v.estado == 'completada')
            ventas_completadas = len([v for v in ventas if v.estado == 'completada'])
            ventas_pendientes = len([v for v in ventas if v.estado == 'pendiente'])
            
            resumen_data = {
                'Resumen de Ventas': [
                    'Total de ventas',
                    'Ventas completadas',
                    'Ventas pendientes',
                    'Ingresos totales (S/)',
                    'Promedio por venta (S/)',
                    'Fecha de generación',
                    'Generado por'
                ],
                'Valor': [
                    len(ventas),
                    ventas_completadas,
                    ventas_pendientes,
                    f"{total_ingresos:.2f}",
                    f"{total_ingresos/ventas_completadas:.2f}" if ventas_completadas > 0 else "0.00",
                    datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                    f"{current_user.nombre} {current_user.apellido}"
                ]
            }
            
            resumen_df = pd.DataFrame(resumen_data)
            resumen_df.to_excel(writer, sheet_name='Resumen', index=False)
            
            # Formatear hoja de resumen
            resumen_sheet = writer.sheets['Resumen']
            for cell in resumen_sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            for column in resumen_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                resumen_sheet.column_dimensions[column_letter].width = adjusted_width
        
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    @app.route('/reportes/pozas/pdf')
    @login_required
    def reporte_pozas_pdf():
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para generar reportes', 'danger')
            return redirect(url_for('dashboard'))
        
        # Obtener datos de pozas
        pozas = Poza.query.all()
        
        # Crear el documento PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            textColor=colors.darkorange,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Contenido del documento
        story = []
        
        # Título
        title = Paragraph("REPORTE DE POZAS - INIA ANDAHUAYLAS", title_style)
        story.append(title)
        
        # Información del reporte
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20
        )
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
        capacidad_total = sum(poza.capacidad for poza in pozas)
        ocupacion_total = sum(len(poza.cuyes) for poza in pozas)
        porcentaje_ocupacion = (ocupacion_total / capacidad_total * 100) if capacidad_total > 0 else 0
        
        info_text = f"<b>Fecha del reporte:</b> {fecha_actual}<br/><b>Total de pozas:</b> {len(pozas)}<br/><b>Capacidad total:</b> {capacidad_total} cuyes<br/><b>Ocupación actual:</b> {ocupacion_total} cuyes ({porcentaje_ocupacion:.1f}%)"
        info_para = Paragraph(info_text, info_style)
        story.append(info_para)
        story.append(Spacer(1, 12))
        
        # Crear tabla de datos
        data = [['ID', 'Código', 'Tipo', 'Capacidad', 'Ocupación', '% Ocupación', 'Estado']]
        
        for poza in pozas:
            cuyes_actuales = len(poza.cuyes)
            ocupacion = (cuyes_actuales / poza.capacidad * 100) if poza.capacidad > 0 else 0
            estado = 'Disponible' if ocupacion < 100 else 'Llena'
            
            data.append([
                str(poza.id),
                poza.codigo,
                poza.tipo.title(),
                str(poza.capacidad),
                str(cuyes_actuales),
                f"{ocupacion:.1f}%",
                estado
            ])
        
        # Crear y estilizar la tabla
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkorange),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Contenido
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Alternar colores de filas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightyellow, colors.white])
        ]))
        
        story.append(table)
        
        # Generar el PDF
        doc.build(story)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_pozas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf',
            mimetype='application/pdf'
        )
    
    @app.route('/reportes/pozas/excel')
    @login_required
    def reporte_pozas_excel():
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para generar reportes', 'danger')
            return redirect(url_for('dashboard'))
        
        # Obtener datos de pozas
        pozas = Poza.query.all()
        
        # Preparar datos para el DataFrame
        data = []
        for poza in pozas:
            cuyes_actuales = len(poza.cuyes)
            ocupacion = (cuyes_actuales / poza.capacidad * 100) if poza.capacidad > 0 else 0
            estado = 'Disponible' if ocupacion < 100 else 'Llena'
            
            data.append({
                'ID': poza.id,
                'Código': poza.codigo,
                'Nombre': poza.nombre or "Sin nombre",
                'Tipo': poza.tipo.title(),
                'Capacidad': poza.capacidad,
                'Ocupación Actual': cuyes_actuales,
                'Porcentaje Ocupación': f"{ocupacion:.1f}%",
                'Estado': estado,
                'Activa': 'Sí' if poza.activa else 'No',
                'Descripción': poza.descripcion or "Sin descripción"
            })
        
        # Crear DataFrame
        df = pd.DataFrame(data)
        
        # Crear buffer para el archivo Excel
        buffer = io.BytesIO()
        
        # Crear el archivo Excel con formato
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Escribir los datos
            df.to_excel(writer, sheet_name='Reporte_Pozas', index=False)
            
            # Obtener el workbook y worksheet para aplicar formato
            workbook = writer.book
            worksheet = writer.sheets['Reporte_Pozas']
            
            # Aplicar formato al encabezado
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Agregar resumen estadístico
            capacidad_total = sum(poza.capacidad for poza in pozas)
            ocupacion_total = sum(len(poza.cuyes) for poza in pozas)
            porcentaje_ocupacion = (ocupacion_total / capacidad_total * 100) if capacidad_total > 0 else 0
            pozas_llenas = len([p for p in pozas if len(p.cuyes) >= p.capacidad])
            pozas_disponibles = len(pozas) - pozas_llenas
            
            resumen_data = {
                'Resumen de Pozas': [
                    'Total de pozas',
                    'Pozas disponibles',
                    'Pozas llenas',
                    'Capacidad total',
                    'Ocupación actual',
                    'Porcentaje ocupación general',
                    'Fecha de generación',
                    'Generado por'
                ],
                'Valor': [
                    len(pozas),
                    pozas_disponibles,
                    pozas_llenas,
                    capacidad_total,
                    ocupacion_total,
                    f"{porcentaje_ocupacion:.1f}%",
                    datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                    f"{current_user.nombre} {current_user.apellido}"
                ]
            }
            
            resumen_df = pd.DataFrame(resumen_data)
            resumen_df.to_excel(writer, sheet_name='Resumen', index=False)
            
            # Formatear hoja de resumen
            resumen_sheet = writer.sheets['Resumen']
            for cell in resumen_sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            for column in resumen_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                resumen_sheet.column_dimensions[column_letter].width = adjusted_width
        
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_pozas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    @app.route('/reportes/estadisticas')
    @login_required
    def estadisticas_generales():
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para ver estadísticas', 'danger')
            return redirect(url_for('dashboard'))
        
        # Estadísticas de cuyes
        total_cuyes = Cuy.query.count()
        cuyes_disponibles = Cuy.query.filter_by(estado='sano').count()
        cuyes_vendidos = Cuy.query.filter_by(estado='vendido').count()
        cuyes_en_tratamiento = Cuy.query.filter_by(estado='en_tratamiento').count()
        
        # Estadísticas de ventas
        total_ventas = Venta.query.count()
        ventas_completadas = Venta.query.filter_by(estado='completada').count()
        ventas_pendientes = Venta.query.filter_by(estado='pendiente').count()
        ventas_rechazadas = Venta.query.filter(Venta.estado.in_(['rechazada', 'cancelada'])).count()
        # Solo considerar ingresos de ventas completadas
        ventas_completadas_list = Venta.query.filter_by(estado='completada').all()
        ingresos_totales = sum(venta.precio for venta in ventas_completadas_list if venta.precio)
        
        # Estadísticas de pozas
        total_pozas = Poza.query.count()
        capacidad_total = sum(poza.capacidad for poza in Poza.query.all())
        # Contar solo cuyes activos en pozas
        ocupacion_total = Cuy.query.filter(Cuy.estado.in_(['sano', 'en_tratamiento']), Cuy.poza_id.isnot(None)).count()
        porcentaje_ocupacion = (ocupacion_total / capacidad_total * 100) if capacidad_total > 0 else 0
        
        # Estadísticas adicionales
        precio_promedio = sum(venta.precio for venta in ventas_completadas_list) / len(ventas_completadas_list) if ventas_completadas_list else 0
        
        content = f"""
        <h2>Estadísticas Generales del Sistema</h2>
        
        <div class='row mb-4'>
            <div class='col-12'>
                <div class='alert alert-primary'>
                    <h4><i class='fas fa-chart-line'></i> Resumen Ejecutivo - INIA Andahuaylas</h4>
                    <p>Fecha del reporte: {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
                    <small><strong>Nota:</strong> Los ingresos incluyen solo ventas completadas</small>
                </div>
            </div>
        </div>
        
        <!-- Estadísticas de Cuyes -->
        <div class='row mb-4'>
            <div class='col-12'>
                <h4><i class='fas fa-chart-pie'></i> Inventario de Cuyes</h4>
            </div>
            <div class='col-md-3 mb-3'>
                <div class='card text-center bg-primary text-white'>
                    <div class='card-body'>
                        <h2>{total_cuyes}</h2>
                        <p>Total de Cuyes</p>
                    </div>
                </div>
            </div>
            <div class='col-md-3 mb-3'>
                <div class='card text-center bg-success text-white'>
                    <div class='card-body'>
                        <h2>{cuyes_disponibles}</h2>
                        <p>Disponibles</p>
                    </div>
                </div>
            </div>
            <div class='col-md-3 mb-3'>
                <div class='card text-center bg-warning text-white'>
                    <div class='card-body'>
                        <h2>{cuyes_en_tratamiento}</h2>
                        <p>En Tratamiento</p>
                    </div>
                </div>
            </div>
            <div class='col-md-3 mb-3'>
                <div class='card text-center bg-info text-white'>
                    <div class='card-body'>
                        <h2>{cuyes_vendidos}</h2>
                        <p>Vendidos</p>
                    </div>
                </div>
            </div>
        </div>
        
        <!-- Estadísticas de Ventas -->
        <div class='row mb-4'>
            <div class='col-12'>
                <h4><i class='fas fa-shopping-cart'></i> Rendimiento de Ventas</h4>
            </div>
            <div class='col-md-2 mb-3'>
                <div class='card text-center bg-secondary text-white'>
                    <div class='card-body'>
                        <h2>{total_ventas}</h2>
                        <p>Total Ventas</p>
                    </div>
                </div>
            </div>
            <div class='col-md-2 mb-3'>
                <div class='card text-center bg-success text-white'>
                    <div class='card-body'>
                        <h2>{ventas_completadas}</h2>
                        <p>Completadas</p>
                    </div>
                </div>
            </div>
            <div class='col-md-2 mb-3'>
                <div class='card text-center bg-warning text-white'>
                    <div class='card-body'>
                        <h2>{ventas_pendientes}</h2>
                        <p>Pendientes</p>
                    </div>
                </div>
            </div>
            <div class='col-md-2 mb-3'>
                <div class='card text-center bg-danger text-white'>
                    <div class='card-body'>
                        <h2>{ventas_rechazadas}</h2>
                        <p>Rechazadas</p>
                    </div>
                </div>
            </div>
            <div class='col-md-4 mb-3'>
                <div class='card text-center bg-success text-white'>
                    <div class='card-body'>
                        <h4>S/ {ingresos_totales:.2f}</h4>
                        <p>Ingresos Totales</p>
                        <small class='text-light'>Precio promedio: S/ {precio_promedio:.2f}</small>
                    </div>
                </div>
            </div>
        </div>
        
        <!-- Estadísticas de Pozas -->
        <div class='row mb-4'>
            <div class='col-12'>
                <h4><i class='fas fa-home'></i> Gestión de Pozas</h4>
            </div>
            <div class='col-md-3 mb-3'>
                <div class='card text-center bg-dark text-white'>
                    <div class='card-body'>
                        <h2>{total_pozas}</h2>
                        <p>Total Pozas</p>
                    </div>
                </div>
            </div>
            <div class='col-md-3 mb-3'>
                <div class='card text-center bg-info text-white'>
                    <div class='card-body'>
                        <h2>{capacidad_total}</h2>
                        <p>Capacidad Total</p>
                    </div>
                </div>
            </div>
            <div class='col-md-3 mb-3'>
                <div class='card text-center bg-primary text-white'>
                    <div class='card-body'>
                        <h2>{ocupacion_total}</h2>
                        <p>Cuyes Alojados</p>
                    </div>
                </div>
            </div>
            <div class='col-md-3 mb-3'>
                <div class='card text-center bg-{'success' if porcentaje_ocupacion < 80 else 'warning' if porcentaje_ocupacion < 100 else 'danger'} text-white'>
                    <div class='card-body'>
                        <h2>{porcentaje_ocupacion:.1f}%</h2>
                        <p>Ocupación</p>
                    </div>
                </div>
            </div>
        </div>
        
        <!-- Distribución por Raza -->
        <div class='row mb-4'>
            <div class='col-12'>
                <h4><i class='fas fa-dna'></i> Distribución por Raza</h4>
                <div class='table-responsive'>
                    <table class='table table-striped'>
                        <thead class='table-dark'>
                            <tr>
                                <th>Raza</th>
                                <th>Cantidad</th>
                                <th>Porcentaje</th>
                            </tr>
                        </thead>
                        <tbody>
        """
        
        # Corregir consulta de razas usando la relación correcta
        razas_stats = db.session.query(Raza.nombre, db.func.count(Cuy.id)).join(Cuy, Raza.id == Cuy.raza_id).group_by(Raza.nombre).all()
        if not razas_stats:
            # Si no hay relaciones, mostrar cuyes sin raza definida
            cuyes_sin_raza = Cuy.query.filter_by(raza_id=None).count()
            if cuyes_sin_raza > 0:
                razas_stats = [('Sin raza definida', cuyes_sin_raza)]
        
        for raza, cantidad in razas_stats:
            porcentaje = (cantidad / total_cuyes * 100) if total_cuyes > 0 else 0
            content += f"""
                            <tr>
                                <td>{raza}</td>
                                <td>{cantidad}</td>
                                <td>{porcentaje:.1f}%</td>
                            </tr>
            """
        
        if not razas_stats:
            content += """
                            <tr>
                                <td colspan='3' class='text-center text-muted'>No hay datos de razas disponibles</td>
                            </tr>
            """
        
        content += """
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
        
        <div class='row'>
            <div class='col-12'>
                <div class='alert alert-success'>
                    <h5><i class='fas fa-lightbulb'></i> Recomendaciones del Sistema</h5>
                    <ul>
        """
        
        if porcentaje_ocupacion > 90:
            content += "<li>⚠️ Pozas cerca del límite de capacidad. Considerar expansión.</li>"
        if cuyes_en_tratamiento > 0:
            content += f"<li>🏥 {cuyes_en_tratamiento} cuyes en tratamiento requieren seguimiento.</li>"
        if ventas_pendientes > 0:
            content += f"<li>📋 {ventas_pendientes} ventas pendientes de procesamiento.</li>"
        if ventas_rechazadas > 0:
            content += f"<li>❌ {ventas_rechazadas} ventas fueron rechazadas. Revisar causas.</li>"
        if total_cuyes > 0:
            porcentaje_disponibles = (cuyes_disponibles / total_cuyes * 100)
            if porcentaje_disponibles > 70:
                content += "<li>✅ Buen inventario disponible para ventas.</li>"
        
        content += """
                    </ul>
                </div>
            </div>
        </div>
        
        <div class='mt-3'>
            <a href='/reportes' class='btn btn-secondary'>Volver a Reportes</a>
            <a href='/reportes/estadisticas/pdf' class='btn btn-primary'>
                <i class='fas fa-file-pdf'></i> Generar PDF
            </a>
        </div>
        """
        
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    @app.route('/reportes/estadisticas/pdf')
    @login_required
    def estadisticas_pdf():
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para generar reportes', 'danger')
            return redirect(url_for('dashboard'))
        
        # Obtener estadísticas (igual que en la función estadisticas_generales)
        total_cuyes = Cuy.query.count()
        cuyes_disponibles = Cuy.query.filter_by(estado='sano').count()
        cuyes_vendidos = Cuy.query.filter_by(estado='vendido').count()
        cuyes_en_tratamiento = Cuy.query.filter_by(estado='en_tratamiento').count()
        
        # Estadísticas de ventas
        total_ventas = Venta.query.count()
        ventas_completadas = Venta.query.filter_by(estado='completada').count()
        ventas_pendientes = Venta.query.filter_by(estado='pendiente').count()
        ventas_rechazadas = Venta.query.filter(Venta.estado.in_(['rechazada', 'cancelada'])).count()
        ventas_completadas_list = Venta.query.filter_by(estado='completada').all()
        ingresos_totales = sum(venta.precio for venta in ventas_completadas_list if venta.precio)
        precio_promedio = sum(venta.precio for venta in ventas_completadas_list) / len(ventas_completadas_list) if ventas_completadas_list else 0
        
        # Estadísticas de pozas
        total_pozas = Poza.query.count()
        capacidad_total = sum(poza.capacidad for poza in Poza.query.all())
        ocupacion_total = Cuy.query.filter(Cuy.estado.in_(['sano', 'en_tratamiento']), Cuy.poza_id.isnot(None)).count()
        porcentaje_ocupacion = (ocupacion_total / capacidad_total * 100) if capacidad_total > 0 else 0
        
        # Crear el documento PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=20,
            textColor=colors.darkblue,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.darkgreen,
            spaceAfter=15,
            spaceBefore=20
        )
        
        # Contenido del documento
        story = []
        
        # Título
        title = Paragraph("ESTADÍSTICAS GENERALES DEL SISTEMA<br/>INIA ANDAHUAYLAS", title_style)
        story.append(title)
        
        # Información del reporte
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20
        )
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
        info_text = f"<b>Fecha del reporte:</b> {fecha_actual}<br/><b>Generado por:</b> {current_user.nombre} {current_user.apellido}<br/><i>Los ingresos incluyen solo ventas completadas</i>"
        info_para = Paragraph(info_text, info_style)
        story.append(info_para)
        story.append(Spacer(1, 20))
        
        # Estadísticas de Cuyes
        cuyes_subtitle = Paragraph("INVENTARIO DE CUYES", subtitle_style)
        story.append(cuyes_subtitle)
        
        cuyes_data = [
            ['Concepto', 'Valor'],
            ['Total de Cuyes', str(total_cuyes)],
            ['Cuyes Disponibles', str(cuyes_disponibles)],
            ['Cuyes en Tratamiento', str(cuyes_en_tratamiento)],
            ['Cuyes Vendidos', str(cuyes_vendidos)]
        ]
        
        cuyes_table = Table(cuyes_data, colWidths=[3*inch, 2*inch])
        cuyes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightblue, colors.white])
        ]))
        story.append(cuyes_table)
        story.append(Spacer(1, 15))
        
        # Estadísticas de Ventas
        ventas_subtitle = Paragraph("RENDIMIENTO DE VENTAS", subtitle_style)
        story.append(ventas_subtitle)
        
        ventas_data = [
            ['Concepto', 'Valor'],
            ['Total de Ventas', str(total_ventas)],
            ['Ventas Completadas', str(ventas_completadas)],
            ['Ventas Pendientes', str(ventas_pendientes)],
            ['Ventas Rechazadas', str(ventas_rechazadas)],
            ['Ingresos Totales', f'S/ {ingresos_totales:.2f}'],
            ['Precio Promedio', f'S/ {precio_promedio:.2f}']
        ]
        
        ventas_table = Table(ventas_data, colWidths=[3*inch, 2*inch])
        ventas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightgreen, colors.white])
        ]))
        story.append(ventas_table)
        story.append(Spacer(1, 15))
        
        # Estadísticas de Pozas
        pozas_subtitle = Paragraph("GESTIÓN DE POZAS", subtitle_style)
        story.append(pozas_subtitle)
        
        pozas_data = [
            ['Concepto', 'Valor'],
            ['Total de Pozas', str(total_pozas)],
            ['Capacidad Total', str(capacidad_total)],
            ['Cuyes Alojados', str(ocupacion_total)],
            ['Porcentaje de Ocupación', f'{porcentaje_ocupacion:.1f}%']
        ]
        
        pozas_table = Table(pozas_data, colWidths=[3*inch, 2*inch])
        pozas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkorange),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightyellow, colors.white])
        ]))
        story.append(pozas_table)
        story.append(Spacer(1, 20))
        
        # Distribución por Raza
        razas_subtitle = Paragraph("DISTRIBUCIÓN POR RAZA", subtitle_style)
        story.append(razas_subtitle)
        
        razas_stats = db.session.query(Raza.nombre, db.func.count(Cuy.id)).join(Cuy, Raza.id == Cuy.raza_id).group_by(Raza.nombre).all()
        if not razas_stats:
            cuyes_sin_raza = Cuy.query.filter_by(raza_id=None).count()
            if cuyes_sin_raza > 0:
                razas_stats = [('Sin raza definida', cuyes_sin_raza)]
        
        if razas_stats:
            razas_data = [['Raza', 'Cantidad', 'Porcentaje']]
            for raza, cantidad in razas_stats:
                porcentaje = (cantidad / total_cuyes * 100) if total_cuyes > 0 else 0
                razas_data.append([raza, str(cantidad), f'{porcentaje:.1f}%'])
        else:
            razas_data = [['Raza', 'Cantidad', 'Porcentaje'], ['No hay datos disponibles', '-', '-']]
        
        razas_table = Table(razas_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
        razas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.purple),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lavender, colors.white])
        ]))
        story.append(razas_table)
        story.append(Spacer(1, 20))
        
        # Recomendaciones
        recomendaciones_subtitle = Paragraph("RECOMENDACIONES DEL SISTEMA", subtitle_style)
        story.append(recomendaciones_subtitle)
        
        recomendaciones = []
        if porcentaje_ocupacion > 90:
            recomendaciones.append("⚠️ Pozas cerca del límite de capacidad. Considerar expansión.")
        if cuyes_en_tratamiento > 0:
            recomendaciones.append(f"🏥 {cuyes_en_tratamiento} cuyes en tratamiento requieren seguimiento.")
        if ventas_pendientes > 0:
            recomendaciones.append(f"📋 {ventas_pendientes} ventas pendientes de procesamiento.")
        if ventas_rechazadas > 0:
            recomendaciones.append(f"❌ {ventas_rechazadas} ventas fueron rechazadas. Revisar causas y mejorar proceso.")
        if total_cuyes > 0:
            porcentaje_disponibles = (cuyes_disponibles / total_cuyes * 100)
            if porcentaje_disponibles > 70:
                recomendaciones.append("✅ Buen inventario disponible para ventas.")
        
        if not recomendaciones:
            recomendaciones.append("✅ Sistema operando dentro de parámetros normales.")
        
        for recomendacion in recomendaciones:
            recom_para = Paragraph(f"• {recomendacion}", styles['Normal'])
            story.append(recom_para)
            story.append(Spacer(1, 8))
        
        # Generar el PDF
        doc.build(story)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'estadisticas_generales_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf',
            mimetype='application/pdf'
        )
    
    @app.route('/logout')
    @login_required
    def logout():
        logout_user()
        return redirect(url_for('index'))
    
    # ========== DASHBOARD MEJORADO CON CONTROL DE ROLES ==========
    @app.route('/dashboard_new')
    @login_required
    def dashboard_new():
        # Obtener estadísticas básicas
        total_cuyes = Cuy.query.count()
        cuyes_disponibles = Cuy.query.filter_by(estado='sano').count()
        total_pozas = Poza.query.count()
        total_ventas = Venta.query.count()
        
        # DASHBOARD PARA ADMINISTRADOR
        if current_user.is_admin():
            ventas_pendientes = Venta.query.filter_by(estado='pendiente').count()
            cuyes_en_tratamiento = Cuy.query.filter_by(estado='en_tratamiento').count()
            total_usuarios = Usuario.query.count()
            
            content = f"""
            <h2><i class='fas fa-tachometer-alt'></i> Dashboard Administrativo</h2>
            
            <div class='alert alert-info'>
                <i class='fas fa-user-shield'></i>
                <strong>Bienvenido Administrador, {current_user.nombre}</strong><br>
                Acceso completo al sistema | {datetime.now().strftime('%d/%m/%Y %H:%M')}
            </div>
            
            <div class='row mb-4'>
                <div class='col-md-3 mb-3'>
                    <div class='card text-center bg-primary text-white'>
                        <div class='card-body'>
                            <i class='fas fa-paw fa-2x mb-2'></i>
                            <h3>{total_cuyes}</h3>
                            <p>Total Cuyes</p>
                        </div>
                    </div>
                </div>
                <div class='col-md-3 mb-3'>
                    <div class='card text-center bg-success text-white'>
                        <div class='card-body'>
                            <i class='fas fa-check-circle fa-2x mb-2'></i>
                            <h3>{cuyes_disponibles}</h3>
                            <p>Disponibles</p>
                        </div>
                    </div>
                </div>
                <div class='col-md-3 mb-3'>
                    <div class='card text-center bg-warning text-white'>
                        <div class='card-body'>
                            <i class='fas fa-exclamation-triangle fa-2x mb-2'></i>
                            <h3>{ventas_pendientes}</h3>
                            <p>Ventas Pendientes</p>
                        </div>
                    </div>
                </div>
                <div class='col-md-3 mb-3'>
                    <div class='card text-center bg-info text-white'>
                        <div class='card-body'>
                            <i class='fas fa-users fa-2x mb-2'></i>
                            <h3>{total_usuarios}</h3>
                            <p>Usuarios</p>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class='row'>
                <div class='col-md-4 mb-3'>
                    <div class='card'>
                        <div class='card-header bg-primary text-white'>
                            <h5><i class='fas fa-paw'></i> Gestión de Cuyes</h5>
                        </div>
                        <div class='card-body'>
                            <p>Administra el inventario completo.</p>
                            <div class='d-grid gap-2'>
                                <a href='/cuyes' class='btn btn-primary'>Ver Cuyes</a>
                                <a href='/cuyes/nuevo' class='btn btn-outline-primary'>Agregar Cuy</a>
                            </div>
                        </div>
                    </div>
                </div>
                
                <div class='col-md-4 mb-3'>
                    <div class='card'>
                        <div class='card-header bg-success text-white'>
                            <h5><i class='fas fa-shopping-cart'></i> Ventas</h5>
                        </div>
                        <div class='card-body'>
                            <p>Gestión de ventas y clientes.</p>
                            <div class='d-grid gap-2'>
                                <a href='/ventas' class='btn btn-success'>Ver Ventas</a>
                                <a href='/catalogo' class='btn btn-outline-success'>Catálogo</a>
                            </div>
                        </div>
                    </div>
                </div>
                
                <div class='col-md-4 mb-3'>
                    <div class='card'>
                        <div class='card-header bg-dark text-white'>
                            <h5><i class='fas fa-cogs'></i> Administración</h5>
                        </div>
                        <div class='card-body'>
                            <p>Configuración del sistema.</p>
                            <div class='d-grid gap-2'>
                                <a href='/usuarios' class='btn btn-dark'>Usuarios</a>
                                <a href='/razas' class='btn btn-outline-dark'>Razas</a>
                                <a href='/pozas' class='btn btn-outline-dark'>Pozas</a>
                            </div>
                        </div>
                    </div>
                </div>
                
                <div class='col-md-4 mb-3'>
                    <div class='card'>
                        <div class='card-header bg-info text-white'>
                            <h5><i class='fas fa-chart-bar'></i> Reportes</h5>
                        </div>
                        <div class='card-body'>
                            <p>Estadísticas y reportes.</p>
                            <div class='d-grid gap-2'>
                                <a href='/reportes' class='btn btn-info'>Ver Reportes</a>
                                <a href='/controles' class='btn btn-outline-info'>Controles</a>
                            </div>
                        </div>
                    </div>
                </div>
                
                <div class='col-md-4 mb-3'>
                    <div class='card'>
                        <div class='card-header bg-warning text-white'>
                            <h5><i class='fas fa-medical-kit'></i> Salud</h5>
                        </div>
                        <div class='card-body'>
                            <p>Control sanitario y tratamientos.</p>
                            <div class='d-grid gap-2'>
                                <a href='/tratamientos' class='btn btn-warning'>Tratamientos</a>
                                <a href='/controles' class='btn btn-outline-warning'>Controles</a>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
            """
            
        # DASHBOARD PARA EMPLEADO  
        elif current_user.is_empleado():
            ventas_pendientes = Venta.query.filter_by(estado='pendiente').count()
            cuyes_en_tratamiento = Cuy.query.filter_by(estado='en_tratamiento').count()
            
            content = f"""
            <h2><i class='fas fa-user-tie'></i> Dashboard Empleado</h2>
            
            <div class='alert alert-success'>
                <i class='fas fa-user-tie'></i>
                <strong>Bienvenido, {current_user.nombre}</strong><br>
                Empleado INIA | {datetime.now().strftime('%d/%m/%Y %H:%M')}
            </div>
            
            <div class='row mb-4'>
                <div class='col-md-4 mb-3'>
                    <div class='card text-center bg-primary text-white'>
                        <div class='card-body'>
                            <i class='fas fa-paw fa-2x mb-2'></i>
                            <h3>{total_cuyes}</h3>
                            <p>Total Cuyes</p>
                        </div>
                    </div>
                </div>
                <div class='col-md-4 mb-3'>
                    <div class='card text-center bg-success text-white'>
                        <div class='card-body'>
                            <i class='fas fa-check-circle fa-2x mb-2'></i>
                            <h3>{cuyes_disponibles}</h3>
                            <p>Disponibles</p>
                        </div>
                    </div>
                </div>
                <div class='col-md-4 mb-3'>
                    <div class='card text-center bg-warning text-white'>
                        <div class='card-body'>
                            <i class='fas fa-exclamation-triangle fa-2x mb-2'></i>
                            <h3>{ventas_pendientes}</h3>
                            <p>Ventas Pendientes</p>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class='row'>
                <div class='col-md-6 mb-3'>
                    <div class='card'>
                        <div class='card-header bg-primary text-white'>
                            <h5><i class='fas fa-paw'></i> Gestión de Cuyes</h5>
                        </div>
                        <div class='card-body'>
                            <p>Administra el inventario de cuyes.</p>
                            <div class='d-grid gap-2'>
                                <a href='/cuyes' class='btn btn-primary'>Ver Cuyes</a>
                            </div>
                        </div>
                    </div>
                </div>
                
                <div class='col-md-6 mb-3'>
                    <div class='card'>
                        <div class='card-header bg-success text-white'>
                            <h5><i class='fas fa-shopping-cart'></i> Ventas</h5>
                        </div>
                        <div class='card-body'>
                            <p>Gestión de ventas.</p>
                            <div class='d-grid gap-2'>
                                <a href='/ventas' class='btn btn-success'>Ver Ventas</a>
                                <a href='/catalogo' class='btn btn-outline-success'>Catálogo</a>
                            </div>
                        </div>
                    </div>
                </div>
                
                <div class='col-md-6 mb-3'>
                    <div class='card'>
                        <div class='card-header bg-info text-white'>
                            <h5><i class='fas fa-chart-bar'></i> Reportes</h5>
                        </div>
                        <div class='card-body'>
                            <p>Reportes básicos.</p>
                            <div class='d-grid gap-2'>
                                <a href='/reportes' class='btn btn-info'>Ver Reportes</a>
                            </div>
                        </div>
                    </div>
                </div>
                
                <div class='col-md-6 mb-3'>
                    <div class='card'>
                        <div class='card-header bg-warning text-white'>
                            <h5><i class='fas fa-medical-kit'></i> Control Sanitario</h5>
                        </div>
                        <div class='card-body'>
                            <p>Tratamientos y controles.</p>
                            <div class='d-grid gap-2'>
                                <a href='/tratamientos' class='btn btn-warning'>Tratamientos</a>
                                <a href='/controles' class='btn btn-outline-warning'>Controles</a>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
            """
            
        # DASHBOARD PARA CLIENTE
        else:
            mis_compras = Venta.query.filter_by(cliente_id=current_user.id).count()
            mis_compras_pendientes = Venta.query.filter_by(cliente_id=current_user.id, estado='pendiente').count()
            
            content = f"""
            <div class='row justify-content-center mb-4'>
                <div class='col-md-10'>
                    <div class='card border-0 shadow-sm'>
                        <div class='card-header bg-gradient text-white text-center' style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);'>
                            <h2 class='mb-0'><i class='fas fa-user-circle'></i> Mi Panel de Cliente</h2>
                            <p class='mb-0 mt-2'>Bienvenido, {current_user.nombre} | Cliente INIA</p>
                        </div>
                        <div class='card-body bg-light'>
                            <div class='row text-center'>
                                <div class='col-md-4 mb-3'>
                                    <div class='card border-0 shadow-sm h-100'>
                                        <div class='card-body bg-white'>
                                            <div class='text-success mb-3'>
                                                <i class='fas fa-paw fa-3x'></i>
                                            </div>
                                            <h3 class='text-success'>{cuyes_disponibles}</h3>
                                            <p class='text-muted'>Cuyes Disponibles</p>
                                        </div>
                                    </div>
                                </div>
                                <div class='col-md-4 mb-3'>
                                    <div class='card border-0 shadow-sm h-100'>
                                        <div class='card-body bg-white'>
                                            <div class='text-primary mb-3'>
                                                <i class='fas fa-shopping-bag fa-3x'></i>
                                            </div>
                                            <h3 class='text-primary'>{mis_compras}</h3>
                                            <p class='text-muted'>Mis Compras</p>
                                        </div>
                                    </div>
                                </div>
                                <div class='col-md-4 mb-3'>
                                    <div class='card border-0 shadow-sm h-100'>
                                        <div class='card-body bg-white'>
                                            <div class='text-warning mb-3'>
                                                <i class='fas fa-clock fa-3x'></i>
                                            </div>
                                            <h3 class='text-warning'>{mis_compras_pendientes}</h3>
                                            <p class='text-muted'>Compras Pendientes</p>
                                        </div>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class='row justify-content-center'>
                <div class='col-md-5 mb-4'>
                    <div class='card h-100 border-0 shadow-sm'>
                        <div class='card-header bg-success text-white text-center'>
                            <h5 class='mb-0'><i class='fas fa-store'></i> Catálogo de Cuyes</h5>
                        </div>
                        <div class='card-body text-center d-flex flex-column'>
                            <p class='flex-grow-1'>Explora nuestro catálogo de cuyes de alta calidad genética disponibles para compra.</p>
                            <a href='/catalogo' class='btn btn-success btn-lg'>
                                <i class='fas fa-eye'></i> Ver Catálogo
                            </a>
                        </div>
                    </div>
                </div>
                
                <div class='col-md-5 mb-4'>
                    <div class='card h-100 border-0 shadow-sm'>
                        <div class='card-header bg-primary text-white text-center'>
                            <h5 class='mb-0'><i class='fas fa-shopping-bag'></i> Mis Compras</h5>
                        </div>
                        <div class='card-body text-center d-flex flex-column'>
                            <p class='flex-grow-1'>Revisa el historial completo de tus compras y el estado de tus pedidos.</p>
                            <a href='/mis-compras' class='btn btn-primary btn-lg'>
                                <i class='fas fa-list'></i> Ver Mis Compras
                            </a>
                        </div>
                    </div>
                </div>
            </div>
            """
        
        content += """
        <div class='mt-4'>
            <div class='alert alert-light border'>
                <h6><i class='fas fa-info-circle text-primary'></i> Sistema de Gestión de Cuyes - INIA Andahuaylas</h6>
                <p class='mb-0'>Plataforma integral para la administración, venta y seguimiento de cuyes de alta calidad genética.</p>
            </div>
        </div>
        """
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    # ========== GESTIÓN DE USUARIOS ==========
    @app.route('/usuarios')
    @login_required
    def usuarios():
        if not current_user.is_admin():
            flash('No tienes permisos para gestionar usuarios', 'danger')
            return redirect(url_for('dashboard'))
        
        usuarios_list = Usuario.query.all()
        
        content = """
        <h2><i class="fas fa-users"></i> Gestión de Usuarios</h2>
        
        <div class="row mb-4">
            <div class="col-12">
                <a href="/usuarios/nuevo" class="btn btn-success">
                    <i class="fas fa-plus"></i> Nuevo Usuario
                </a>
            </div>
        </div>
        
        <div class="row">
            <div class="col-12">
                <div class="card">
                    <div class="card-header">
                        <h5><i class="fas fa-list"></i> Lista de Usuarios</h5>
                    </div>
                    <div class="card-body">
                        <div class="table-responsive">
                            <table class="table table-striped">
                                <thead>
                                    <tr>
                                        <th>ID</th>
                                        <th>Nombre</th>
                                        <th>Email</th>
                                        <th>Rol</th>
                                        <th>Fecha Registro</th>
                                        <th>Estado</th>
                                        <th>Acciones</th>
                                    </tr>
                                </thead>
                                <tbody>
        """
        
        for usuario in usuarios_list:
            estado_badge = "success" if usuario.activo else "danger"
            estado_texto = "Activo" if usuario.activo else "Inactivo"
            rol_badge = {
                'admin': 'danger',
                'empleado': 'warning', 
                'cliente': 'info'
            }.get(usuario.rol, 'secondary')
            
            content += f"""
                                    <tr>
                                        <td>{usuario.id}</td>
                                        <td>{usuario.nombre_completo}</td>
                                        <td>{usuario.email}</td>
                                        <td><span class="badge bg-{rol_badge}">{usuario.rol.title()}</span></td>
                                        <td>{usuario.fecha_registro.strftime('%d/%m/%Y') if usuario.fecha_registro else 'N/A'}</td>
                                        <td><span class="badge bg-{estado_badge}">{estado_texto}</span></td>
                                        <td>
                                            <a href="/usuarios/{usuario.id}/editar" class="btn btn-sm btn-primary">
                                                <i class="fas fa-edit"></i>
                                            </a>
                                            <a href="/usuarios/{usuario.id}/toggle" class="btn btn-sm btn-{'danger' if usuario.activo else 'success'}" 
                                               onclick="return confirm('¿Estás seguro?')">
                                                <i class="fas fa-{'ban' if usuario.activo else 'check'}"></i>
                                            </a>
                                        </td>
                                    </tr>
            """
        
        content += """
                                </tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        """
        
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    @app.route('/usuarios/nuevo', methods=['GET', 'POST'])
    @login_required
    def nuevo_usuario():
        if not current_user.is_admin():
            flash('No tienes permisos para crear usuarios', 'danger')
            return redirect(url_for('dashboard'))
        
        if request.method == 'POST':
            # Validar que el email no exista
            email = request.form.get('email')
            if Usuario.query.filter_by(email=email).first():
                flash('El email ya está registrado', 'danger')
                return redirect(url_for('nuevo_usuario'))
            
            usuario = Usuario(
                nombre=request.form.get('nombre'),
                apellido=request.form.get('apellido'),
                email=email,
                password_hash=generate_password_hash(request.form.get('password')),
                rol=request.form.get('rol'),
                activo=True
            )
            
            db.session.add(usuario)
            db.session.commit()
            
            flash('Usuario creado exitosamente', 'success')
            return redirect(url_for('usuarios'))
        
        form_content = """
        <h2><i class="fas fa-user-plus"></i> Nuevo Usuario</h2>
        
        <div class="row justify-content-center">
            <div class="col-md-6">
                <div class="card">
                    <div class="card-body">
                        <form method="POST">
                            <div class="mb-3">
                                <label for="nombre" class="form-label">Nombre</label>
                                <input type="text" class="form-control" id="nombre" name="nombre" required>
                            </div>
                            
                            <div class="mb-3">
                                <label for="apellido" class="form-label">Apellido</label>
                                <input type="text" class="form-control" id="apellido" name="apellido" required>
                            </div>
                            
                            <div class="mb-3">
                                <label for="email" class="form-label">Email</label>
                                <input type="email" class="form-control" id="email" name="email" required>
                            </div>
                            
                            <div class="mb-3">
                                <label for="password" class="form-label">Contraseña</label>
                                <input type="password" class="form-control" id="password" name="password" required>
                            </div>
                            
                            <div class="mb-3">
                                <label for="rol" class="form-label">Rol</label>
                                <select class="form-control" id="rol" name="rol" required>
                                    <option value="">Seleccionar rol...</option>
                                    <option value="admin">Administrador</option>
                                    <option value="empleado">Empleado</option>
                                    <option value="cliente">Cliente</option>
                                </select>
                            </div>
                            
                            <div class="d-grid gap-2">
                                <button type="submit" class="btn btn-success">
                                    <i class="fas fa-save"></i> Crear Usuario
                                </button>
                                <a href="/usuarios" class="btn btn-secondary">
                                    <i class="fas fa-arrow-left"></i> Volver
                                </a>
                            </div>
                        </form>
                    </div>
                </div>
            </div>
        </div>
        """
        
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', form_content))
    
    @app.route('/usuarios/<int:user_id>/editar', methods=['GET', 'POST'])
    @login_required
    def editar_usuario(user_id):
        if not current_user.is_admin():
            flash('No tienes permisos para editar usuarios', 'danger')
            return redirect(url_for('dashboard'))
        
        usuario = Usuario.query.get_or_404(user_id)
        
        if request.method == 'POST':
            # Validar que el email no exista en otro usuario
            email = request.form.get('email')
            existing_user = Usuario.query.filter_by(email=email).first()
            if existing_user and existing_user.id != user_id:
                flash('El email ya está registrado por otro usuario', 'danger')
                return redirect(url_for('editar_usuario', user_id=user_id))
            
            usuario.nombre = request.form.get('nombre')
            usuario.apellido = request.form.get('apellido')
            usuario.email = email
            usuario.rol = request.form.get('rol')
            
            # Solo cambiar contraseña si se proporciona una nueva
            new_password = request.form.get('password')
            if new_password:
                usuario.password_hash = generate_password_hash(new_password)
            
            db.session.commit()
            
            flash('Usuario actualizado exitosamente', 'success')
            return redirect(url_for('usuarios'))
        
        form_content = f"""
        <h2><i class="fas fa-edit"></i> Editar Usuario</h2>
        
        <div class="row justify-content-center">
            <div class="col-md-6">
                <div class="card">
                    <div class="card-body">
                        <form method="POST">
                            <div class="mb-3">
                                <label for="nombre" class="form-label">Nombre</label>
                                <input type="text" class="form-control" id="nombre" name="nombre" value="{usuario.nombre}" required>
                            </div>
                            
                            <div class="mb-3">
                                <label for="apellido" class="form-label">Apellido</label>
                                <input type="text" class="form-control" id="apellido" name="apellido" value="{usuario.apellido}" required>
                            </div>
                            
                            <div class="mb-3">
                                <label for="email" class="form-label">Email</label>
                                <input type="email" class="form-control" id="email" name="email" value="{usuario.email}" required>
                            </div>
                            
                            <div class="mb-3">
                                <label for="password" class="form-label">Nueva Contraseña (dejar vacío para mantener actual)</label>
                                <input type="password" class="form-control" id="password" name="password">
                            </div>
                            
                            <div class="mb-3">
                                <label for="rol" class="form-label">Rol</label>
                                <select class="form-control" id="rol" name="rol" required>
                                    <option value="admin" {'selected' if usuario.rol == 'admin' else ''}>Administrador</option>
                                    <option value="empleado" {'selected' if usuario.rol == 'empleado' else ''}>Empleado</option>
                                    <option value="cliente" {'selected' if usuario.rol == 'cliente' else ''}>Cliente</option>
                                </select>
                            </div>
                            
                            <div class="d-grid gap-2">
                                <button type="submit" class="btn btn-primary">
                                    <i class="fas fa-save"></i> Actualizar Usuario
                                </button>
                                <a href="/usuarios" class="btn btn-secondary">
                                    <i class="fas fa-arrow-left"></i> Volver
                                </a>
                            </div>
                        </form>
                    </div>
                </div>
            </div>
        </div>
        """
        
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', form_content))
    
    @app.route('/usuarios/<int:user_id>/toggle')
    @login_required
    def toggle_usuario(user_id):
        if not current_user.is_admin():
            flash('No tienes permisos para cambiar el estado de usuarios', 'danger')
            return redirect(url_for('dashboard'))
        
        usuario = Usuario.query.get_or_404(user_id)
        
        # No permitir desactivar al propio admin
        if usuario.id == current_user.id and usuario.rol == 'admin':
            flash('No puedes desactivar tu propia cuenta de administrador', 'warning')
            return redirect(url_for('usuarios'))
        
        usuario.activo = not usuario.activo
        db.session.commit()
        
        estado = "activado" if usuario.activo else "desactivado"
        flash(f'Usuario {estado} exitosamente', 'success')
        
        return redirect(url_for('usuarios'))
    
    # Inicializar base de datos y datos básicos
    with app.app_context():
        try:
            db.create_all()
            
            # Crear datos básicos si no existen
            if not Usuario.query.filter_by(email='admin@inia.gob.pe').first():
                admin = Usuario(
                    nombre='Administrador',
                    apellido='Sistema',
                    email='admin@inia.gob.pe',
                    password_hash=generate_password_hash('admin123'),
                    rol='admin'
                )
                db.session.add(admin)
                
                # Usuarios de ejemplo
                empleado = Usuario(
                    nombre='Juan',
                    apellido='Pérez',
                    email='empleado@inia.gob.pe',
                    password_hash=generate_password_hash('empleado123'),
                    rol='empleado'
                )
                db.session.add(empleado)
                
                cliente = Usuario(
                    nombre='María',
                    apellido='González',
                    email='cliente@example.com',
                    password_hash=generate_password_hash('cliente123'),
                    rol='cliente'
                )
                db.session.add(cliente)
                
                # Razas básicas
                razas = [
                    Raza(nombre='Perú', descripcion='Raza criolla peruana'),
                    Raza(nombre='Andina', descripcion='Raza mejorada andina'),
                    Raza(nombre='Inti', descripcion='Línea INIA')
                ]
                for raza in razas:
                    db.session.add(raza)
                
                # Pozas básicas
                pozas = [
                    Poza(codigo='P001', tipo='reproductores', capacidad=20),
                    Poza(codigo='P002', tipo='lactantes', capacidad=30),
                    Poza(codigo='P003', tipo='recria', capacidad=50)
                ]
                for poza in pozas:
                    db.session.add(poza)
                
                db.session.commit()
                print("✓ Datos básicos creados")
            else:
                print("✓ Datos básicos ya existen")
            
            # Actualizar ocupación de pozas al iniciar
            actualizar_ocupacion_pozas()
            print("✓ Ocupación de pozas actualizada")
                
        except Exception as e:
            print(f"⚠ Error de base de datos: {e}")
    
    print("✓ Servidor completo iniciado")
    print("📧 Credenciales: admin@inia.gob.pe / admin123")
    print("🌐 URL: http://localhost:5000")
    
    # ========== RUTAS ADICIONALES ==========
    @app.route('/admin/ventas')
    @login_required
    def admin_ventas():
        # Solo admin y empleado pueden acceder
        if not (current_user.is_admin() or current_user.is_empleado()):
            flash('No tienes permisos para acceder a esta sección', 'danger')
            return redirect(url_for('dashboard'))
        
        # Obtener todas las ventas con información de cliente y cuy
        ventas = db.session.query(Venta).join(Usuario, Venta.cliente_id == Usuario.id).join(Cuy).all()
        
        content = f"""
        <div class='d-flex justify-content-between align-items-center mb-4'>
            <h2><i class='fas fa-chart-line'></i> Gestión de Ventas</h2>
            <div>
                <span class='badge bg-info me-2'>Total: {len(ventas)}</span>
                <span class='badge bg-warning me-2'>Pendientes: {len([v for v in ventas if v.estado == 'pendiente'])}</span>
                <span class='badge bg-success'>Completadas: {len([v for v in ventas if v.estado == 'completada'])}</span>
            </div>
        </div>
        
        <div class='card'>
            <div class='card-body'>
                <div class='table-responsive'>
                    <table class='table table-striped'>
                        <thead class='table-dark'>
                            <tr>
                                <th>ID</th>
                                <th>Fecha</th>
                                <th>Cliente</th>
                                <th>Cuy</th>
                                <th>Precio</th>
                                <th>Estado</th>
                                <th>Acciones</th>
                            </tr>
                        </thead>
                        <tbody>
        """
        
        for venta in ventas:
            fecha = venta.fecha_venta.strftime('%d/%m/%Y') if venta.fecha_venta else 'N/A'
            cliente_nombre = f"{venta.cliente.nombre} {venta.cliente.apellido}" if venta.cliente else 'N/A'
            cuy_codigo = venta.cuy.codigo if venta.cuy else 'N/A'
            
            estado_class = {
                'pendiente': 'warning',
                'completada': 'success',
                'cancelada': 'danger'
            }.get(venta.estado, 'secondary')
            
            content += f"""
                            <tr>
                                <td>#{venta.id}</td>
                                <td>{fecha}</td>
                                <td>{cliente_nombre}</td>
                                <td>{cuy_codigo}</td>
                                <td>S/ {venta.precio:.2f}</td>
                                <td><span class='badge bg-{estado_class}'>{venta.estado.title()}</span></td>
                                <td>
            """
            
            if venta.estado == 'pendiente':
                content += f"""
                                    <form method="POST" action="/venta/{venta.id}/completar" style="display:inline;">
                                        <button type="submit" class="btn btn-sm btn-success me-1" onclick="return confirm('¿Marcar como completada?')">
                                            <i class='fas fa-check'></i> Completar
                                        </button>
                                    </form>
                                    <form method="POST" action="/venta/{venta.id}/cancelar" style="display:inline;">
                                        <button type="submit" class="btn btn-sm btn-danger" onclick="return confirm('¿Cancelar venta?')">
                                            <i class='fas fa-times'></i> Cancelar
                                        </button>
                                    </form>
                """
            else:
                content += f"""
                                    <span class='text-muted'>
                                        <i class='fas fa-lock'></i> {venta.estado.title()}
                                    </span>
                """
            
            content += """
                                </td>
                            </tr>
            """
        
        if not ventas:
            content += """
                            <tr>
                                <td colspan='7' class='text-center text-muted'>
                                    <i class='fas fa-info-circle'></i> No hay ventas registradas
                                </td>
                            </tr>
            """
        
        content += """
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
        
        <div class='mt-3'>
            <a href='/dashboard' class='btn btn-secondary'>
                <i class='fas fa-arrow-left'></i> Volver al Dashboard
            </a>
        </div>
        """
        
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    @app.route('/perfil')
    @login_required
    def perfil():
        content = f"""
        <div class='row justify-content-center'>
            <div class='col-md-8'>
                <div class='card'>
                    <div class='card-header bg-primary text-white'>
                        <h4><i class='fas fa-user'></i> Mi Perfil</h4>
                    </div>
                    <div class='card-body'>
                        <div class='row'>
                            <div class='col-md-6'>
                                <p><strong>Nombre:</strong> {current_user.nombre}</p>
                                <p><strong>Apellido:</strong> {current_user.apellido}</p>
                                <p><strong>Email:</strong> {current_user.email}</p>
                            </div>
                            <div class='col-md-6'>
                                <p><strong>Rol:</strong> <span class='badge bg-info'>{current_user.rol.title()}</span></p>
                                <p><strong>Teléfono:</strong> {current_user.telefono or 'No registrado'}</p>
                                <p><strong>Estado:</strong> <span class='badge bg-success'>Activo</span></p>
                            </div>
                        </div>
                        
                        {f"<p><strong>Dirección:</strong> {current_user.direccion}</p>" if current_user.direccion else ""}
                        
                        <hr>
                        <div class='d-grid gap-2 d-md-flex justify-content-md-end'>
                            <a href='/perfil/editar' class='btn btn-primary'>
                                <i class='fas fa-edit'></i> Editar Perfil
                            </a>
                            <a href='/dashboard' class='btn btn-secondary'>
                                <i class='fas fa-arrow-left'></i> Volver
                            </a>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        """
        
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))
    
    @app.route('/perfil/editar', methods=['GET', 'POST'])
    @login_required
    def editar_perfil():
        if request.method == 'POST':
            # Actualizar datos del perfil
            current_user.nombre = request.form.get('nombre', current_user.nombre)
            current_user.apellido = request.form.get('apellido', current_user.apellido)
            current_user.telefono = request.form.get('telefono', current_user.telefono)
            current_user.direccion = request.form.get('direccion', current_user.direccion)
            
            # Cambiar contraseña si se proporciona
            nueva_password = request.form.get('nueva_password')
            if nueva_password and len(nueva_password) >= 6:
                current_user.password_hash = generate_password_hash(nueva_password)
            
            db.session.commit()
            flash('Perfil actualizado correctamente', 'success')
            return redirect(url_for('perfil'))
        
        content = f"""
        <div class='row justify-content-center'>
            <div class='col-md-8'>
                <div class='card'>
                    <div class='card-header bg-primary text-white'>
                        <h4><i class='fas fa-user-edit'></i> Editar Perfil</h4>
                    </div>
                    <div class='card-body'>
                        <form method='POST'>
                            <div class='row'>
                                <div class='col-md-6 mb-3'>
                                    <label for='nombre' class='form-label'>Nombre *</label>
                                    <input type='text' class='form-control' id='nombre' name='nombre' 
                                           value='{current_user.nombre}' required>
                                </div>
                                <div class='col-md-6 mb-3'>
                                    <label for='apellido' class='form-label'>Apellido *</label>
                                    <input type='text' class='form-control' id='apellido' name='apellido' 
                                           value='{current_user.apellido}' required>
                                </div>
                            </div>
                            
                            <div class='mb-3'>
                                <label for='email' class='form-label'>Email</label>
                                <input type='email' class='form-control' id='email' value='{current_user.email}' disabled>
                                <div class='form-text'>El email no puede ser modificado</div>
                            </div>
                            
                            <div class='row'>
                                <div class='col-md-6 mb-3'>
                                    <label for='telefono' class='form-label'>Teléfono</label>
                                    <input type='tel' class='form-control' id='telefono' name='telefono' 
                                           value='{current_user.telefono or ""}'>
                                </div>
                                <div class='col-md-6 mb-3'>
                                    <label for='rol' class='form-label'>Rol</label>
                                    <input type='text' class='form-control' value='{current_user.rol.title()}' disabled>
                                </div>
                            </div>
                            
                            <div class='mb-3'>
                                <label for='direccion' class='form-label'>Dirección</label>
                                <textarea class='form-control' id='direccion' name='direccion' rows='2'>{current_user.direccion or ""}</textarea>
                            </div>
                            
                            <hr>
                            <h5>Cambiar Contraseña (Opcional)</h5>
                            <div class='mb-3'>
                                <label for='nueva_password' class='form-label'>Nueva Contraseña</label>
                                <input type='password' class='form-control' id='nueva_password' name='nueva_password' 
                                       minlength='6' placeholder='Dejar vacío si no desea cambiar'>
                                <div class='form-text'>Mínimo 6 caracteres</div>
                            </div>
                            
                            <div class='d-grid gap-2 d-md-flex justify-content-md-end'>
                                <a href='/perfil' class='btn btn-secondary me-md-2'>
                                    <i class='fas fa-times'></i> Cancelar
                                </a>
                                <button type='submit' class='btn btn-primary'>
                                    <i class='fas fa-save'></i> Guardar Cambios
                                </button>
                            </div>
                        </form>
                    </div>
                </div>
            </div>
        </div>
        """
        
        return render_template_string(BASE_TEMPLATE.replace('{{content}}', content))

    # Ejecutar servidor
    app.run(debug=True, host='127.0.0.1', port=5000)
    
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()
